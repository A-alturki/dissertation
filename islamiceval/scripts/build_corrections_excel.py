#!/usr/bin/env python3
"""
build_corrections_excel.py
==========================
Complete correction-review spreadsheet from the one-shot pipeline, GROUNDED against
the corpus (with --partial --alef_deletion --snap as the "after" setting).

Per-span rows (one row per span):
    id | model | span | ref | gpt verdict | gemini verdict | verdict agreement |
    gpt grounded corrections | gemini grounded corrections |
    correction agreement (normalized exact, on grounded) | # matches (when yes)

Then a STATISTICS block at the bottom:
  1. correctable / not-correctable rate per correction model, overall + per ref.
  2. corpus-match % per model, overall + per ref, BEFORE vs AFTER (--partial --alef_deletion --snap).
  3. verdict overlap between the two models (both-ok / both-خطأ / disagree).
  4. of spans where both models are ok, % with >=1 matching grounded correction.

"grounded" = exact-substring grounded OR snapped (>=0.90). The grounded-correction
TEXT compared is the corpus-canonical (snapped -> gold text; exact -> matched entry).
Agreement normalization = un-diacritized + space-stripped + alif-insensitive.

Usage:
    python build_corrections_excel.py
    python build_corrections_excel.py --input ../outputs/correction/pipeline_oneshot.json \
                                      --out   ../outputs/correction/corrections_complete.xlsx
"""
import os, json, argparse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

import ground_corrections as G
import corpus_search as CS          # deterministic top-K retriever (for manual inspection column)

GPT, GEM = "gpt-5.4", "gemini-3.1-pro-preview"
SNAP_TH = 0.90
RETRIEVE_K = 3                      # top-K corpus_search candidates shown per span
RETRIEVE_COVERAGE = 0.20           # drop a candidate whose run covers < this fraction of the span words
RETRIEVE_TEXT_CAP = 300            # cap each retrieved candidate's displayed text (chars)
PART = False        # whether the "after" grounding uses --partial (set from CLI)
SNAP_ON = True      # whether the "after" grounding includes snapping (set from CLI)
AFTER_LABEL = ""    # human label for the after setting, set in main()

# agreement normalization: un-diacritized + space-stripped + alif-insensitive
def anorm(t):
    if not t:
        return ""
    return G._nonletter.sub("", t.translate(G._TABLE)).replace("ا", "")


def build_corpora():
    """Strict corpus (alif kept), alef corpus (alif dropped), snap corpus (matn-only)."""
    G.ALEF_DEL = False
    strict = G.load_corpus()
    G.ALEF_DEL = True
    alef = G.load_corpus()
    snapc = G.load_snap_corpus()
    return strict, alef, snapc


def ground_before(text, ref, strict):
    G.ALEF_DEL = False; G.PARTIAL = False
    ok, *_ = G.lookup(text, ref, *strict)
    return ok

def ground_after(text, ref, claimed, alef, snapc):
    """(grounded_bool, resolve_dict|None) for the stats 'after' column. In snapped mode
    this is the layered resolve (Quran->full ayah, Hadith snap[+expand], ref fallback,
    else eliminated); in --no_snap mode it's exact-only alef lookup."""
    G.ALEF_DEL = True; G.PARTIAL = PART
    if not SNAP_ON:
        ok, *_ = G.lookup(text, ref, *alef)
        return ok, None
    r = G.resolve_correction(text, ref, claimed, *alef, snapc)
    return r["status"] != "eliminated", r


def resolve_model(block, model, ref, alef, snapc):
    """(grounded, eliminated) for a model's ok candidates.
      grounded   = [(gold_text, sources_list, status)]   (status: grounded|snapped|ref)
                   — one entry PER distinct expanded variant (hadith expansion); the
                   exact-duplicate collapse across books happens in fmt_corrections.
      eliminated = [(model_text, claimed_source)]"""
    p = ((block or {}).get(model) or {}).get("prediction") or {}
    if p.get("verdict") != "ok":
        return [], []
    grounded, elim = [], []
    G.ALEF_DEL = True; G.PARTIAL = PART
    for c in (p.get("candidates") or []):
        r = G.resolve_correction(c.get("text", ""), ref, c.get("source", ""), *alef, snapc)
        if r["status"] == "eliminated":
            elim.append((c.get("text", ""), c.get("source", "")))
        else:
            for vtext, vsources in (r.get("variants") or [(r["gold_text"], r["sources"])]):
                grounded.append((vtext, vsources, r["status"]))
    return grounded, elim


def raw_model(block, model):
    """Raw, unfiltered model corrections as (text, [claimed_source], 'raw') — reference view."""
    p = ((block or {}).get(model) or {}).get("prediction") or {}
    if p.get("verdict") != "ok":
        return []
    return [(c.get("text", ""), [c.get("source", "")], "raw") for c in (p.get("candidates") or [])]


def fmt_corrections(items, elim):
    """Collapse ONLY exact-identical matns (un-diacritized + space-stripped, ALIF-KEPT)
    across books into one line with sources unioned; alif-distinct variants stay on
    separate lines. Then list eliminated candidates, clearly tagged."""
    by, order = {}, []
    for text, sources, _ in items:
        k = G.norm_dedup(text)
        if not k:
            continue
        if k not in by:
            by[k] = [text, list(sources)]; order.append(k)
        else:
            for sset in sources:
                if sset not in by[k][1]:
                    by[k][1].append(sset)
    lines = [f"{by[k][0]}  [{'؛ '.join(by[k][1])}]" for k in order]
    lines += [f"ELIMINATED (no corpus match) [{cs}]: {t}" for t, cs in elim]
    return "\n".join(lines)


def verdict_of(block, m):
    p = ((block or {}).get(m) or {}).get("prediction")
    return "ERR" if p is None else p.get("verdict", "?")


def retrieved_refs(span, ref, k=RETRIEVE_K):
    """Deterministic top-K corpus_search candidates for the span (source + run + text)
    for manual inspection. Drops weak hits (run covering < RETRIEVE_COVERAGE of the
    span's words, so 2-3-word generic matches on long stories disappear) and caps each
    candidate's text at RETRIEVE_TEXT_CAP chars (the matched_run line carries the key
    phrase, so long matns don't blow up the cell)."""
    try:
        cands = CS.search(span, ref, k=max(k * 3, 10))     # over-fetch, then filter
    except Exception as e:
        return f"(search error: {e})"
    nsw = len(CS.norm(span).split()) or 1
    strong = [c for c in cands if c.get("run_len", 0) / nsw >= RETRIEVE_COVERAGE][:k]
    if not strong:
        return "(no strong corpus candidates)"
    out = []
    for i, c in enumerate(strong, 1):
        txt = c.get("text", "")
        if len(txt) > RETRIEVE_TEXT_CAP:
            txt = txt[:RETRIEVE_TEXT_CAP].rstrip() + " …"
        cov = c.get("run_len", 0) / nsw
        out.append(f"[{i}] {c.get('source','')}  (run {c.get('run_len','?')}, "
                   f"ov {c.get('overlap','?')}, cov {cov:.0%})\n"
                   f"    run: {c.get('matched_run','')}\n"
                   f"    {txt}")
    return "\n".join(out)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", default=os.path.join(G.HERE, "..", "outputs", "correction", "pipeline_oneshot.json"))
    ap.add_argument("--out",   default=None,
                    help="output xlsx; defaults to corrections_complete[_nosnap].xlsx by --no_snap")
    ap.add_argument("--partial", action="store_true",
                    help="include --partial (reverse-substring) in the 'after' grounding setting")
    ap.add_argument("--no_snap", action="store_true",
                    help="disable snapping: 'after' grounding = exact-only (still --alef_deletion)")
    ap.add_argument("--hadith_expansion", action="store_true",
                    help="for each snapped hadith, expand to ALL corpus matns >=95% similar or "
                         "substring-equal (the full multi-book attestation set), single-hop")
    args = ap.parse_args()

    global PART, SNAP_ON, AFTER_LABEL
    PART = args.partial
    SNAP_ON = not args.no_snap
    G.HADITH_EXPAND = args.hadith_expansion
    AFTER_LABEL = ("--alef_deletion" + (" --snap" if SNAP_ON else "") + (" --partial" if PART else "")
                   + (" --hadith_expansion" if args.hadith_expansion else ""))
    if args.out is None:
        name = ("corrections_complete" + ("" if SNAP_ON else "_nosnap")
                + ("_expanded" if args.hadith_expansion else "") + ".xlsx")
        args.out = os.path.join(G.HERE, "..", "outputs", "correction", name)

    print(f"loading corpora (strict + alef + snap); after = {AFTER_LABEL}", flush=True)
    strict, alef, snapc = build_corpora()
    data = json.load(open(args.input, encoding="utf-8"))

    rows = []          # display rows
    recs = []          # structured per-span for stats
    for s in data["per_sample"]:
        osh, ref = s.get("oneshot", {}), s["ref"]
        vg, vge = verdict_of(osh, GPT), verdict_of(osh, GEM)
        if SNAP_ON:                                        # grounded/snapped canonical corrections
            cg,  elim_g = resolve_model(osh, GPT, ref, alef, snapc)
            cge, elim_e = resolve_model(osh, GEM, ref, alef, snapc)
        else:                                              # raw model outputs, unfiltered
            cg, cge, elim_g, elim_e = raw_model(osh, GPT), raw_model(osh, GEM), [], []
        # agreement on the shown correction texts (normalized exact)
        sg  = {anorm(t) for t, _, _ in cg if anorm(t)}
        sge = {anorm(t) for t, _, _ in cge if anorm(t)}
        if "ERR" in (vg, vge):
            agree, nmatch = "—", ""
        elif vg == "خطأ" and vge == "خطأ":
            agree, nmatch = "both خطأ", ""
        elif vg != vge:
            agree, nmatch = "no", ""
        else:                                       # both ok
            inter = sg & sge
            agree = "yes" if inter else "no"
            nmatch = len(inter) if inter else 0

        # per-candidate grounding for stats (before / after), per model
        def cand_ground(model):
            p = ((osh or {}).get(model) or {}).get("prediction") or {}
            if p.get("verdict") != "ok":
                return []
            res = []
            for c in (p.get("candidates") or []):
                tx = c.get("text", "")
                b = ground_before(tx, ref, strict)
                a, _ = ground_after(tx, ref, c.get("source", ""), alef, snapc)
                res.append((b, a))
            return res

        rec = {"id": s["id"], "ref": ref, "vg": vg, "vge": vge,
               "agree": agree, "nmatch": nmatch,
               "gpt_grounded": cg, "gem_grounded": cge,
               "gpt_cand": cand_ground(GPT), "gem_cand": cand_ground(GEM)}
        recs.append(rec)
        rows.append([
            s["id"], s.get("answer_model", ""), s["span"], ref, vg, vge,
            "yes" if vg == vge else "no",
            fmt_corrections(cg, elim_g),
            fmt_corrections(cge, elim_e),
            agree, nmatch,
            retrieved_refs(s["span"], ref),
        ])

    write_excel(args.out, rows, recs)
    print(f"  -> wrote {args.out}")


# ─────────────────────────────────────────────────────────────────────────────
def _rate(recs, model_key, ref=None):
    """(ok, total) verdicts for a correction model, optionally filtered by ref."""
    vk = "vg" if model_key == GPT else "vge"
    sub = [r for r in recs if (ref is None or r["ref"] == ref) and r[vk] in ("ok", "خطأ")]
    ok = sum(1 for r in sub if r[vk] == "ok")
    return ok, len(sub)

def _corpus_match(recs, cand_key, ref=None):
    """(grounded_before, grounded_after, total_candidates) for a model's candidates."""
    b = a = n = 0
    for r in recs:
        if ref is not None and r["ref"] != ref:
            continue
        for gb, ga in r[cand_key]:
            n += 1; b += gb; a += ga
    return b, a, n


def write_excel(path, rows, recs):
    corr_label = ("corrections (raw model output)" if not SNAP_ON
                  else "grounded corrections (Quran=full ayah; Hadith"
                       + (" expanded)" if G.HADITH_EXPAND else ")"))
    HEAD = ["id", "model", "span", "ref", "gpt verdict", "gemini verdict",
            "verdict agreement", f"gpt {corr_label}", f"gemini {corr_label}",
            "correction agreement (normalized exact)", "# matches",
            f"retrieved refs (corpus_search top-{RETRIEVE_K}: source | run | text)"]
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "corrections"
    ws.sheet_view.rightToLeft = True
    hfill, hfont = PatternFill("solid", fgColor="305496"), Font(bold=True, color="FFFFFF")
    green, orange = PatternFill("solid", fgColor="E2EFDA"), PatternFill("solid", fgColor="FCE4D6")

    for j, h in enumerate(HEAD, 1):
        c = ws.cell(1, j, h); c.fill, c.font = hfill, hfont
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ri, row in enumerate(rows, 2):
        for j, v in enumerate(row, 1):
            c = ws.cell(ri, j, v)
            c.alignment = Alignment(vertical="top", wrap_text=True,
                                    horizontal="right" if j in (3, 8, 9, 12) else "center")
        ws.cell(ri, 7).fill  = green if row[6] == "yes" else orange
        ws.cell(ri, 10).fill = green if row[9] in ("yes", "both خطأ") else orange
    for j, w in enumerate([9, 13, 52, 7, 10, 12, 11, 50, 50, 16, 9, 70], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"

    # ── STATISTICS block ──
    r = len(rows) + 4
    def title(txt):
        nonlocal r
        c = ws.cell(r, 1, txt); c.font = Font(bold=True, size=12); r += 1
    def hdr(cells):
        nonlocal r
        for j, t in enumerate(cells, 1):
            c = ws.cell(r, j, t); c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="D9E1F2")
        r += 1
    def line(cells):
        nonlocal r
        for j, t in enumerate(cells, 1):
            ws.cell(r, j, t)
        r += 1

    # 1. correctable rate
    title("1. Correctable rate per model (verdict = ok), overall + per ref")
    hdr(["model", "scope", "correctable", "not (خطأ)", "total", "correctable %"])
    for mk, name in [(GPT, "gpt-5.4"), (GEM, "gemini-3.1-pro")]:
        for ref in (None, "quran", "hadith"):
            ok, tot = _rate(recs, mk, ref)
            line([name, ref or "overall", ok, tot - ok, tot, f"{100*ok/tot:.1f}%" if tot else "—"])
    r += 1

    # 2. corpus-match % before/after
    title(f"2. Corpus-match % per model (of candidates), before (strict exact) vs after ({AFTER_LABEL})")
    hdr(["model", "scope", "candidates", "before %", "after %"])
    for mk, ck, name in [(GPT, "gpt_cand", "gpt-5.4"), (GEM, "gem_cand", "gemini-3.1-pro")]:
        for ref in (None, "quran", "hadith"):
            b, a, n = _corpus_match(recs, ck, ref)
            line([name, ref or "overall", n,
                  f"{100*b/n:.1f}%" if n else "—", f"{100*a/n:.1f}%" if n else "—"])
    r += 1

    # 3. verdict overlap between models
    title("3. Verdict overlap between the two models")
    both_ok  = sum(1 for x in recs if x["vg"] == "ok"  and x["vge"] == "ok")
    both_kh  = sum(1 for x in recs if x["vg"] == "خطأ" and x["vge"] == "خطأ")
    valid    = [x for x in recs if x["vg"] in ("ok", "خطأ") and x["vge"] in ("ok", "خطأ")]
    disagree = len(valid) - both_ok - both_kh
    hdr(["both correctable", "both not (خطأ)", "disagree", "agreement %", "n"])
    line([both_ok, both_kh, disagree,
          f"{100*(both_ok+both_kh)/len(valid):.1f}%" if valid else "—", len(valid)])
    r += 1

    # 4. of both-ok spans, % with >=1 matching grounded correction
    title("4. Of spans where BOTH models are ok: % with >=1 matching grounded correction")
    both_ok_rows = [x for x in recs if x["vg"] == "ok" and x["vge"] == "ok"]
    with_match = sum(1 for x in both_ok_rows if x["agree"] == "yes")
    hdr(["both-ok spans", "with >=1 grounded text match", "%"])
    line([len(both_ok_rows), with_match,
          f"{100*with_match/len(both_ok_rows):.1f}%" if both_ok_rows else "—"])

    d = os.path.dirname(path)
    if d:
        os.makedirs(d, exist_ok=True)
    wb.save(path)


if __name__ == "__main__":
    main()
