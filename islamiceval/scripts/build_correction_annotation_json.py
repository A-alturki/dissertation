#!/usr/bin/env python3
"""
build_correction_annotation_json.py
===================================
Assemble the single JSON the correction annotation tool reads. One object per curated
span (from sample_correction_final_classified.csv), carrying everything the annotator
needs:

  id, ref, model (the answer model), error_type,
  question  (prompt, from the annotated sample sheet),
  answer    (full model answer, from the annotated sample sheet),
  span + span_start/span_end  (offset of the span inside `answer`, for highlighting),
  correctors.first  = gpt-5.4               {verdict, corrections:[{text, sources}]}
  correctors.second = gemini-3.1-pro-preview {verdict, corrections:[{text, sources}]}

The corrections are the GROUNDED + HADITH-EXPANDED ones (same layered policy as
build_corrections_excel: Quran -> full ayah; Hadith snap + expand to every corpus
matn variant; collapse only EXACT un-diacritized [alif-kept] duplicates).

Sources:
  - curated set:      ../outputs/correction/sample_correction_final_classified.csv
  - gpt/gemini preds: ../outputs/correction/pipeline_oneshot.json  (oneshot block)
  - question+answer:  ../outputs/answers/sample65_span_correctness_annotation/<model>_sample_annotated.xlsx

Usage:
    python build_correction_annotation_json.py
    python build_correction_annotation_json.py --out ../outputs/correction/correction_annotation.json
"""
import os, sys, csv, json, argparse
from datetime import datetime, timezone
from openpyxl import load_workbook

import ground_corrections as G
import build_corrections_excel as BCE

HERE   = os.path.dirname(os.path.abspath(__file__))
CORR   = os.path.join(HERE, "..", "outputs", "correction")
# the curated set is now TWO batches: the original 100 + the 2026-06-23 batch2 sourced
# from the colleague's minor_issues/no_match JSONL. Both merge into one 200-item JSON.
# batch2 answers/questions are NOT in the sample sheets — they come from batch2_context.json
# (id,model,span -> question, answer, offsets recovered from outputs/answers/<set>/).
CSVS    = [os.path.join(CORR, "sample_correction_final_classified.csv"),
           os.path.join(CORR, "sample_correction_batch2.csv")]
ONESHOTS= [os.path.join(CORR, "pipeline_oneshot.json"),
           os.path.join(CORR, "pipeline_oneshot_batch2.json")]
CONTEXTS= [os.path.join(CORR, "batch2_context.json")]
SHEETS = os.path.join(HERE, "..", "outputs", "answers", "sample65_span_correctness_annotation")

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass


def load_sheet(model):
    """(id) -> (prompt, answer) from <model>_sample_annotated.xlsx."""
    path = os.path.join(SHEETS, f"{model}_sample_annotated.xlsx")
    wb = load_workbook(path, read_only=True, data_only=True)
    it = wb.active.iter_rows(values_only=True)
    hdr = list(next(it))
    qi, ai, ii = hdr.index("prompt"), hdr.index("answer"), hdr.index("id")
    out = {}
    for r in it:
        if r[ii] is not None:
            out[r[ii]] = (r[qi] or "", r[ai] or "")
    return out


def corrections_for(block, model, ref, alef, snapc):
    """Grounded + hadith-expanded corrections for one corrector, deduped by exact
    (un-diacritized, alif-kept) text with sources unioned — mirrors fmt_corrections."""
    grounded, _elim = BCE.resolve_model(block, model, ref, alef, snapc)
    by, order = {}, []
    for text, sources, _status in grounded:
        k = G.norm_dedup(text)
        if not k:
            continue
        if k not in by:
            by[k] = {"text": text, "sources": list(sources)}
            order.append(k)
        else:
            for s in sources:
                if s not in by[k]["sources"]:
                    by[k]["sources"].append(s)
    return [by[k] for k in order]


def verdict_of(block, model):
    p = ((block or {}).get(model) or {}).get("prediction") or {}
    return p.get("verdict", "?")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv", nargs="+", default=CSVS, help="curated span CSV(s), merged in order")
    ap.add_argument("--oneshot", nargs="+", default=ONESHOTS, help="pipeline oneshot JSON(s), merged")
    ap.add_argument("--context", nargs="*", default=CONTEXTS,
                    help="context JSON(s) supplying question/answer/offsets for spans not in the sheets")
    ap.add_argument("--out", default=os.path.join(CORR, "correction_annotation.json"))
    args = ap.parse_args()

    # grounding settings = the "lenient expansion" setup (alef + snap + partial + hadith_expansion)
    G.HADITH_EXPAND = True
    BCE.PART = True
    BCE.SNAP_ON = True
    print("loading corpora (strict + alef + snap)...", flush=True)
    _strict, alef, snapc = BCE.build_corpora()

    rows = []
    for path in args.csv:
        if not os.path.exists(path):
            print(f"  WARNING: CSV not found, skipping: {path}")
            continue
        rows.extend(csv.DictReader(open(path, encoding="utf-8-sig")))
    oneshot = {}
    for path in args.oneshot:
        if not os.path.exists(path):
            print(f"  WARNING: oneshot JSON not found, skipping: {path}")
            continue
        for s in json.load(open(path, encoding="utf-8"))["per_sample"]:
            oneshot[(s["id"], s["answer_model"], s["span"])] = s
    context = {}         # (id, model, span) -> {question, answer, span_start, span_end}
    for path in (args.context or []):
        if not os.path.exists(path):
            print(f"  WARNING: context JSON not found, skipping: {path}")
            continue
        for c in json.load(open(path, encoding="utf-8"))["items"]:
            context[(c["id"], c["model"], c["span"])] = c
    print(f"  merged {len(rows)} spans from {len(args.csv)} CSV(s); "
          f"{len(oneshot)} oneshot predictions from {len(args.oneshot)} file(s); "
          f"{len(context)} context records")

    sheets = {}          # model -> id->(prompt,answer) cache
    items, miss_answer, miss_pred = [], [], []
    for r in rows:
        mid, model, span, ref = r["id"], r["model"], r["span"], r["ref"]
        ctx = context.get((mid, model, span))
        if ctx is not None:                         # batch2: answer/question from context file
            prompt, answer = ctx["question"], ctx["answer"]
        else:                                       # original 100: from the sample sheets
            if model not in sheets:
                sheets[model] = load_sheet(model)
            prompt, answer = sheets[model].get(mid, ("", ""))
        if not answer or span not in answer:
            miss_answer.append((mid, model))
        s = oneshot.get((mid, model, span))
        if s is None:
            miss_pred.append((mid, model))
        block = (s or {}).get("oneshot", {})
        start = answer.find(span)
        items.append({
            "id": mid, "ref": ref, "model": model,
            "error_type": r.get("error_type", ""),
            "question": prompt, "answer": answer,
            "span": span,
            "span_start": start, "span_end": (start + len(span)) if start >= 0 else -1,
            # corrector identities are deliberately NOT included — the annotator must
            # stay blind to which model (or that it is a model at all) produced each.
            # Mapping for the researcher: first = gpt-5.4, second = gemini-3.1-pro-preview.
            "correctors": {
                "first":  {"verdict": verdict_of(block, BCE.GPT),
                           "corrections": corrections_for(block, BCE.GPT, ref, alef, snapc)},
                "second": {"verdict": verdict_of(block, BCE.GEM),
                           "corrections": corrections_for(block, BCE.GEM, ref, alef, snapc)},
            },
        })

    out = {
        "generated": datetime.now(timezone.utc).isoformat(),
        "n": len(items),
        "grounding": "alef_deletion + snap + partial + hadith_expansion",
        "items": items,
    }
    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    json.dump(out, open(args.out, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    print(f"  wrote {len(items)} items -> {args.out}")
    if miss_answer:
        print(f"  WARNING: {len(miss_answer)} spans not found in their answer: {miss_answer[:8]}")
    if miss_pred:
        print(f"  WARNING: {len(miss_pred)} spans missing oneshot prediction: {miss_pred[:8]}")


if __name__ == "__main__":
    main()
