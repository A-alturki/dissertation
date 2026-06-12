"""
Classify Islamic questions on 3 axes with GPT-4.1-nano:
  - category (one of 8), difficulty (1-5), divergence (1-5).

Reads an .xlsx (qid + prompt) or .json, sends each question to the model
concurrently (fast throughput on 8k), enforces structured JSON output, and writes
incrementally with resume. One request per question (the prompt is written for a
single question); concurrency provides the "batched" speed.

Usage:
    python classify_questions.py                                  # default input/output
    python classify_questions.py --input ../data/classified/accepted_combined_fixed.xlsx
    python classify_questions.py --limit 5 --workers 4            # quick test
"""
import os, sys, json, argparse, time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dotenv import load_dotenv
from openai import OpenAI
from tqdm.auto import tqdm

load_dotenv()

DEFAULT_MODEL = "gpt-4.1-nano"
REASONING_PREFIXES = ("gpt-5", "o1", "o3", "o4", "o5")

# USD per 1M tokens: (input, cached_input, output)
PRICES = {
    "gpt-4.1-nano": (0.10, 0.025, 0.40),
    "gpt-5.4-mini": (0.75, 0.075, 4.50),
    "gpt-5.4-nano": (0.20, 0.02,  1.25),
}

CATEGORIES = ["Aqidah", "Fiqh", "Hadith Studies", "Quranic Studies",
              "Sirah and Islamic History", "Inheritance", "Family Law", "Islamic Finance"]

SCHEMA = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "category":   {"type": "string", "enum": CATEGORIES},
        "difficulty": {"type": "integer", "enum": [1, 2, 3, 4, 5]},
        "divergence": {"type": "integer", "enum": [1, 2, 3, 4, 5]},
    },
    "required": ["category", "difficulty", "divergence"],
}

SYSTEM_PROMPT = """You are an expert annotator of Islamic questions. For each question you receive,
you will assign three labels: (1) a Category, (2) a Difficulty score from 1 to 5,
and (3) a Scholarly Divergence score from 1 to 5. Apply the definitions and rules
below precisely. Output JSON only, with no additional text.

═══════════════════════════════════════════════════════════════════
PART 1 — CATEGORY (choose exactly ONE)
═══════════════════════════════════════════════════════════════════

Assign the single category that best matches the PRIMARY subject the question is
testing. If a question touches multiple areas, choose the category of the main
thing being asked, not incidental mentions.

1. "Aqidah" — Islamic creed and belief. Questions about Allah and His attributes,
   prophethood, angels, revealed scriptures, the Day of Judgment, divine decree
   (qadar), faith, disbelief, and theological doctrine.

2. "Fiqh" — Practical rulings on worship and conduct. Questions about purification,
   prayer, fasting, hajj, halal/haram acts, and how to correctly perform a
   religious act or whether something is permitted. This is the default for
   ruling-seeking questions that are not specifically about family, inheritance,
   or finance (which have their own categories below).

3. "Hadith Studies" — The prophetic traditions AS TRADITIONS. Questions about a
   hadith's authentication, grading (sahih/hasan/da'if), narrators, isnad, or the
   science of hadith. NOTE: If a hadith is merely USED to derive or support a
   ruling, classify as the relevant ruling category (usually Fiqh), not here.
   Classify here only when the hadith itself, its chain, or its authenticity is
   the subject.

4. "Quranic Studies" — The Qur'an as text. Questions about verses, their meaning,
   tafsir, themes, Qur'anic stories, occasions of revelation, recitation, and
   interpretation. NOTE: If a verse is merely cited to support a ruling, classify
   as the relevant ruling category. Classify here when the verse or its meaning is
   the subject.

5. "Sirah and Islamic History" — Historical narrative. The life of Prophet
   Muhammad and his companions, and Islamic historical events, figures, dynasties,
   and civilizations across all eras.

6. "Inheritance" — Islamic inheritance (mirath/fara'id). Questions about heirs,
   fixed shares, distribution, and inheritance calculation.

7. "Family Law" — Marriage, divorce, child custody, spousal and family rights,
   and related personal-status rulings.

8. "Islamic Finance" — Halal and haram transactions, banking, riba, business
   contracts, investment, and modern financial products judged by Islamic law.


═══════════════════════════════════════════════════════════════════
PART 2 — DIFFICULTY (integer 1–5)
═══════════════════════════════════════════════════════════════════

Difficulty measures the COGNITIVE EFFORT required to answer the question
responsibly and correctly — how much recall, synthesis, and reasoning it takes.
It does NOT measure how controversial the question is (that is Part 3). A question
can be difficult but uncontroversial (e.g. a multi-step inheritance computation
with one correct answer).

1 — Trivial recall. A single, well-known fact retrievable directly. No reasoning.
    e.g. "How many obligatory prayers are there in a day?"
    e.g. "Who is the prophet of islam?"

2 — Simple recall or a single definition/distinction. One concept, lightly
    explained, no chaining of steps.
    e.g. "What is the difference between wajib and sunnah?"
    e.g. "What does the word 'taqwa' mean?"

3 — Single-step application. Requires applying one rule or concept to the specifics
    of the question, or combining two related facts.
    e.g. "If I forget a rak'ah in prayer, how do I perform sujud al-sahw?"
    e.g. "Does touching one's spouse invalidate wudu?"

4 — Multi-step reasoning. Requires chaining several rules, conditions, or facts
    together, or resolving dependencies before reaching the answer.
    e.g. "Compute the inheritance shares when the deceased leaves a wife, two
    daughters, and both parents."
    e.g. "How does combining and shortening prayers work for a traveler who
    intends to stay nine days?"

5 — Complex synthesis. Requires integrating multiple sources, balancing competing
    principles or evidences, handling many conditions or edge cases, or reasoning
    through a novel scenario with no direct precedent.
    e.g. "Evaluate the permissibility of a conventional mortgage for a Muslim
    minority living where no Islamic financing exists, weighing necessity against
    the prohibition of riba."

Scoring guidance:
- Judge difficulty by what a RESPONSIBLE, CORRECT answer requires — not by how
  short a careless answer could be.
- More required qualifications, conditions, or steps → higher difficulty.
- If the question is underspecified such that answering well requires laying out
  multiple cases, score the effort of handling those cases.

═══════════════════════════════════════════════════════════════════
PART 3 — SCHOLARLY DIVERGENCE (integer 1–5)
═══════════════════════════════════════════════════════════════════

Scholarly Divergence measures the extent to which QUALIFIED SCHOLARS LEGITIMATELY
DIFFER on the answer — across madhahib (schools of law), across valid
interpretations, or across competing evidences that yield different conclusions.
It does NOT measure cognitive effort (that is Part 2) and does NOT measure whether
the topic is socially sensitive. A question can be easy yet highly divergent (a
one-line ruling that the four madhahib answer differently), or hard yet
convergent (a complex computation with one agreed result).

1 — Settled / unanimous. A single agreed answer; effectively no scholarly
    disagreement (ijma' or near-universal agreement).
    e.g. "Is the dawn (Fajr) prayer two rak'ahs?"
    e.g. "Is wine prohibited in Islam?"

2 — Minor variation. Broad agreement on the core answer with only small or
    technical differences in detail or wording among scholars.
    e.g. "What are the conditions of a valid ablution?" (largely agreed, minor
    differences at the margins)

3 — Recognized difference of opinion. Two or more well-established positions exist,
    but one is clearly majority or the differences are modest in practical effect.
    e.g. "Does touching one's spouse invalidate wudu?" (madhahib differ; positions
    are well-known)

4 — Substantial madhhab/interpretive split. Multiple major schools or interpretive
    traditions reach genuinely different rulings, with no single dominant consensus,
    and the practical consequences differ.
    e.g. "Is it permissible to combine prayers without travel or excuse?"
    e.g. "What is the ruling on music?"

5 — Deeply contested / no stable consensus. The answer varies fundamentally across
    schools, eras, or contemporary scholarly bodies; or it is a modern issue on
    which authoritative scholars actively disagree with no settled resolution.
    e.g. "What is the Islamic ruling on cryptocurrency trading?"
    e.g. "Is cosmetic surgery for non-medical reasons permissible?"

Scoring guidance:
- Score by the ANSWER SPACE, not the question's tone. Ask: would qualified scholars
  give materially different rulings?
- Pure factual/historical questions (e.g. dates, who-did-what) are almost always
  1–2: facts are not subject to legal disagreement even when details are debated.
- Modern/novel scenarios with no direct precedent tend toward 4–5.
- If you are unsure whether a genuine scholarly split exists, do not inflate the
  score — reserve 4–5 for cases where you can identify the differing positions.

═══════════════════════════════════════════════════════════════════
OUTPUT FORMAT
═══════════════════════════════════════════════════════════════════

Return ONLY a JSON object, no surrounding text, in exactly this form:

{"category": "<one category string from Part 1>", "difficulty": <integer 1-5>, "divergence": <integer 1-5>}"""


SYSTEM_PROMPT_ARABIC = """
أنت خبير في تصنيف الأسئلة الإسلامية. لكل سؤال تستلمه، سوف تعيّن
ثلاثة تصنيفات: (1) الفئة، (2) درجة الصعوبة من 1 إلى 5،
و(3) درجة الاختلاف الفقهي من 1 إلى 5. طبّق التعريفات والقواعد
أدناه بدقة. أخرج JSON فقط، بدون أي نص إضافي.

═══════════════════════════════════════════════════════════════════
الجزء 1 — الفئة (اختر فئة واحدة فقط)
═══════════════════════════════════════════════════════════════════

عيّن الفئة الوحيدة التي تطابق الموضوع الأساسي الذي يختبره السؤال.
إذا كان السؤال يتطرق لعدة مجالات، اختر فئة الشيء الرئيسي المسؤول
عنه، لا الإشارات العابرة.

1. "Aqidah" — العقيدة الإسلامية. أسئلة عن الله وصفاته،
   والنبوة، والملائكة، والكتب المنزّلة، واليوم الآخر، والقدر،
   والإيمان، والكفر، والمسائل العقدية.

2. "Fiqh" — الأحكام العملية في العبادات والمعاملات. أسئلة عن الطهارة،
   والصلاة، والصيام، والحج، والحلال والحرام، وكيفية أداء عبادة
   بشكل صحيح أو حُكم جواز شيء. هذه هي الفئة الافتراضية للأسئلة
   التي تطلب حكماً ولا تخص الأسرة أو الميراث أو المعاملات المالية
   (ولكلٍّ منها فئتها أدناه).

3. "Hadith Studies" — الأحاديث النبوية بوصفها أحاديث. أسئلة عن
   تخريج الحديث، ودرجته (صحيح/حسن/ضعيف)، ورواته، وإسناده، وعلم
   الحديث. ملاحظة: إذا استُعمل الحديث لمجرد استنباط حكم أو دعمه،
   صنّفه ضمن فئة الحكم المناسبة (غالباً Fiqh)، لا هنا. صنّف هنا
   فقط عندما يكون الحديث نفسه أو سنده أو صحته هو الموضوع.

4. "Quranic Studies" — القرآن بوصفه نصاً. أسئلة عن الآيات،
   ومعانيها، والتفسير، والموضوعات، والقصص القرآني، وأسباب النزول،
   والتلاوة، والتأويل. ملاحظة: إذا استُشهد بآية لمجرد دعم حكم،
   صنّفه ضمن فئة الحكم المناسبة. صنّف هنا عندما تكون الآية أو
   معناها هي الموضوع.

5. "Sirah and Islamic History" — السرد التاريخي. سيرة النبي
   محمد ﷺ وصحابته، والأحداث والشخصيات والدول والحضارات
   الإسلامية عبر العصور.

6. "Inheritance" — المواريث (الميراث/الفرائض). أسئلة عن الورثة،
   والأنصبة المقدّرة، والتوزيع، وحساب الميراث.

7. "Family Law" — النكاح، والطلاق، وحضانة الأطفال، وحقوق
   الزوجين والأسرة، وما يتعلق بأحكام الأحوال الشخصية.

8. "Islamic Finance" — المعاملات الحلال والحرام، والمصارف،
   والربا، والعقود التجارية، والاستثمار، والمنتجات المالية
   الحديثة من منظور الشريعة.


═══════════════════════════════════════════════════════════════════
الجزء 2 — الصعوبة (عدد صحيح 1–5)
═══════════════════════════════════════════════════════════════════

تقيس الصعوبة الجهد الذهني المطلوب للإجابة على السؤال بمسؤولية
وصحة — مقدار الاستحضار والتركيب والاستدلال الذي يتطلبه. وهي لا
تقيس مدى إثارة السؤال للخلاف (ذلك هو الجزء 3). فقد يكون السؤال
صعباً لكنه غير خلافي (مثل مسألة ميراث متعددة الخطوات لها إجابة
واحدة صحيحة).

1 — استحضار بسيط جداً. معلومة واحدة معروفة تُستحضر مباشرة. بلا استدلال.
    مثال: "كم عدد الصلوات المفروضة في اليوم؟"
    مثال: "منهو نبي الإسلام؟"

2 — استحضار بسيط أو تعريف/تمييز واحد. مفهوم واحد بشرح خفيف،
    بلا تسلسل خطوات.
    مثال: "وش الفرق بين الواجب والسنة؟"
    مثال: "وش معنى كلمة 'تقوى'؟"

3 — تطبيق بخطوة واحدة. يتطلب تطبيق قاعدة أو مفهوم واحد على
    تفاصيل السؤال، أو الجمع بين معلومتين مترابطتين.
    مثال: "إذا نسيت ركعة في الصلاة، كيف أسوي سجود السهو؟"
    مثال: "هل لمس الزوجة ينقض الوضوء؟"

4 — استدلال متعدد الخطوات. يتطلب ربط عدة قواعد أو شروط أو حقائق
    معاً، أو حل تبعيات قبل الوصول للجواب.
    مثال: "احسب أنصبة الميراث إذا ترك المتوفى زوجة وبنتين
    والأبوين."
    مثال: "كيف يجمع المسافر ويقصر الصلاة إذا نوى الإقامة تسعة
    أيام؟"

5 — تركيب معقّد. يتطلب دمج عدة مصادر، أو الموازنة بين مبادئ أو
    أدلة متعارضة، أو معالجة شروط وحالات حدّية كثيرة، أو الاستدلال
    في سيناريو جديد بلا سابقة مباشرة.
    مثال: "قيّم حكم القرض العقاري التقليدي لمسلم يعيش كأقلية في
    مكان لا يتوفر فيه تمويل إسلامي، موازناً بين الضرورة وتحريم
    الربا."

إرشادات التقييم:
- قدّر الصعوبة بما يتطلبه جواب مسؤول وصحيح — لا بمدى قِصَر
  جواب متهاون.
- كلما زادت القيود أو الشروط أو الخطوات المطلوبة → ارتفعت الصعوبة.
- إذا كان السؤال ناقص التحديد بحيث يتطلب الجواب الجيد عرض عدة
  حالات، قدّر جهد معالجة تلك الحالات.

═══════════════════════════════════════════════════════════════════
الجزء 3 — الاختلاف الفقهي (عدد صحيح 1–5)
═══════════════════════════════════════════════════════════════════

يقيس الاختلاف الفقهي مدى اختلاف العلماء المؤهّلين اختلافاً
مشروعاً على الجواب — بين المذاهب، أو بين التأويلات المعتبرة، أو
بين الأدلة المتعارضة التي تؤدي إلى استنتاجات مختلفة. وهو لا يقيس
الجهد الذهني (ذلك هو الجزء 2)، ولا يقيس مدى حساسية الموضوع
اجتماعياً. فقد يكون السؤال سهلاً لكنه شديد الخلاف (حكم من سطر
واحد تجيب عليه المذاهب الأربعة بشكل مختلف)، أو صعباً لكنه متفق
عليه (مسألة حسابية معقّدة لها نتيجة واحدة متفق عليها).

1 — مستقر/مُجمَع عليه. جواب واحد متفق عليه؛ لا خلاف فقهي فعلياً
    (إجماع أو شبه اتفاق تام).
    مثال: "هل صلاة الفجر ركعتان؟"
    مثال: "هل الخمر محرّمة في الإسلام؟"

2 — تباين طفيف. اتفاق واسع على أصل الجواب مع اختلافات صغيرة أو
    تقنية في التفصيل أو الصياغة بين العلماء.
    مثال: "وش شروط الوضوء الصحيح؟" (متفق عليها إلى حد كبير، مع
    اختلافات يسيرة في الأطراف)

3 — خلاف معتبر. توجد قولان أو أكثر راسخان، لكن أحدهما أرجح
    بوضوح أو أن الفروق محدودة الأثر العملي.
    مثال: "هل لمس الزوجة ينقض الوضوء؟" (تختلف المذاهب؛ والأقوال
    معروفة)

4 — خلاف مذهبي/تأويلي جوهري. تصل عدة مذاهب أو مدارس تأويلية إلى
    أحكام مختلفة فعلاً، بلا إجماع مهيمن واحد، وتختلف نتائجها
    العملية.
    مثال: "هل يجوز الجمع بين الصلاتين بلا سفر أو عذر؟"
    مثال: "وش حكم الموسيقى؟"

5 — خلاف عميق/بلا إجماع مستقر. يختلف الجواب اختلافاً جوهرياً بين
    المذاهب أو العصور أو الهيئات العلمية المعاصرة؛ أو هو مسألة
    حديثة يختلف فيها العلماء المعتبرون بلا حل مستقر.
    مثال: "وش حكم تداول العملات الرقمية في الإسلام؟"
    مثال: "هل تجوز الجراحة التجميلية لغير غرض طبي؟"

إرشادات التقييم:
- قيّم بناءً على فضاء الإجابات، لا على نبرة السؤال. اسأل: هل
  يعطي العلماء المؤهّلون أحكاماً مختلفة فعلاً؟
- الأسئلة الواقعية/التاريخية البحتة (مثل التواريخ، ومن فعل ماذا)
  تكون غالباً 1–2: فالحقائق لا تخضع للخلاف الفقهي حتى لو نوقشت
  تفاصيلها.
- السيناريوهات الحديثة/المستجدة بلا سابقة مباشرة تميل إلى 4–5.
- إذا لم تكن متأكداً من وجود خلاف فقهي حقيقي، فلا تضخّم الدرجة
  — احفظ 4–5 للحالات التي تستطيع فيها تحديد الأقوال المختلفة.

═══════════════════════════════════════════════════════════════════
صيغة الإخراج
═══════════════════════════════════════════════════════════════════

أخرج كائن JSON فقط، بدون أي نص محيط، بهذه الصيغة تماماً:

{"category": "<فئة واحدة من الجزء 1>", "difficulty": <عدد صحيح 1-5>, "divergence": <عدد صحيح 1-5>}

"""

def load_prompts(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        from openpyxl import load_workbook
        rows = load_workbook(path, read_only=True, data_only=True).active.iter_rows(values_only=True)
        header = list(next(rows))
        qi, pi = header.index("qid"), header.index("prompt")
        return [{"id": r[qi], "prompt": r[pi]} for r in rows if r[pi]]
    if ext == ".json":
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    raise ValueError(f"Unsupported input '{ext}' (use .xlsx or .json)")


def classify_one(client, item, model, is_reasoning):
    """Return (id, record, usage). Retries transient failures; records an error after 5 tries.
    usage = {"in","cached","out"} token counts (None on failure)."""
    for attempt in range(5):
        try:
            kw = dict(
                model=model,
                messages=[{"role": "system", "content": SYSTEM_PROMPT_ARABIC},
                          {"role": "user", "content": str(item["prompt"])}],
                response_format={"type": "json_schema",
                                 "json_schema": {"name": "labels", "strict": True, "schema": SCHEMA}},
            )
            if is_reasoning:
                # gpt-5.x lowest effort is "none"; o-series only down to "low".
                kw["max_completion_tokens"] = 2048
                kw["reasoning_effort"] = "none" if model.startswith("gpt-5") else "low"
            else:
                kw["temperature"] = 0
                kw["max_tokens"] = 60
            r = client.chat.completions.create(**kw)
            obj = json.loads(r.choices[0].message.content)
            u = r.usage
            ptd = getattr(u, "prompt_tokens_details", None)
            cached = (getattr(ptd, "cached_tokens", 0) or 0) if ptd else 0
            usage = {"in": u.prompt_tokens, "cached": cached, "out": u.completion_tokens}
            return item["id"], {"id": item["id"], "prompt": item["prompt"], **obj}, usage
        except Exception as e:
            if attempt == 4:
                return item["id"], {"id": item["id"], "prompt": item["prompt"], "error": str(e)[:200]}, None
            time.sleep(2 ** attempt)


def load_ids(path):
    """Read qids from a json (list of {id,...}) — e.g. a sample file or a labels file."""
    d = json.load(open(path, encoding="utf-8"))
    return {x["id"] for x in d if isinstance(x, dict) and "id" in x}


def main():
    ap = argparse.ArgumentParser(description="Classify Islamic questions (category/difficulty/divergence)")
    ap.add_argument("--model", default=DEFAULT_MODEL, help="OpenAI model id (e.g. gpt-4.1-nano, gpt-5.4-mini, gpt-5.4-nano)")
    ap.add_argument("--input",  default="../data/classified/accepted_combined_fixed.xlsx")
    ap.add_argument("--output", default=None, help="defaults to ../outputs/classification/labels_<model><tag>.json")
    ap.add_argument("--ids-from", default=None, help="restrict to the qids found in this json (sample/labels file)")
    ap.add_argument("--tag", default="", help="suffix for the default output filename (e.g. _100, _sample65)")
    ap.add_argument("--workers", type=int, default=16)
    ap.add_argument("--save-every", type=int, default=100)
    ap.add_argument("--limit", type=int, default=None, help="classify only the first N (testing)")
    ap.add_argument("--api-key", default=None, help="pass the OpenAI key explicitly (overrides .env)")
    args = ap.parse_args()

    key = args.api_key or os.environ.get("OPENAI_API_KEY")
    if not key:
        sys.exit("No OpenAI key — use --api-key, or set OPENAI_API_KEY in .env")

    model = args.model
    is_reasoning = model.startswith(REASONING_PREFIXES)
    output = args.output or f"../outputs/classification/labels_{model}{args.tag}.json"

    prompts = load_prompts(args.input)
    if args.ids_from:
        keep = load_ids(args.ids_from)
        prompts = [p for p in prompts if p["id"] in keep]
        print(f"Restricted to {len(prompts)} questions (ids from {args.ids_from})")
    if args.limit:
        prompts = prompts[:args.limit]
    print(f"Loaded {len(prompts)} questions | model={model} (reasoning={is_reasoning}) -> {output}")

    os.makedirs(os.path.dirname(output), exist_ok=True)
    results = {}
    if os.path.exists(output):
        for r in json.load(open(output, encoding="utf-8")):
            results[r["id"]] = r
    todo = [p for p in prompts if "category" not in results.get(p["id"], {})]
    print(f"{len(results)} already labeled, {len(todo)} to do.")

    def save():
        tmp = output + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(list(results.values()), f, ensure_ascii=False, indent=2)
        os.replace(tmp, output)

    client = OpenAI(api_key=key)
    done = 0
    tin = tcached = tout = 0
    t0 = time.time()
    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        futs = [ex.submit(classify_one, client, p, model, is_reasoning) for p in todo]
        for f in tqdm(as_completed(futs), total=len(futs), desc=model):
            i, rec, usage = f.result()
            results[i] = rec
            if usage:
                tin += usage["in"]; tcached += usage["cached"]; tout += usage["out"]
            done += 1
            if done % args.save_every == 0:
                save()
    save()
    wall = time.time() - t0

    # summary
    ok = [r for r in results.values() if "category" in r]
    err = len(results) - len(ok)
    from collections import Counter
    cats = Counter(r["category"] for r in ok)
    print(f"\nLabeled {len(ok)} ({err} errors) -> {output}")
    print("Category distribution:")
    for c, n in cats.most_common():
        print(f"  {c:28} {n}")
    if ok:
        print(f"Difficulty mean: {sum(r['difficulty'] for r in ok)/len(ok):.2f} | "
              f"Divergence mean: {sum(r['divergence'] for r in ok)/len(ok):.2f}")

    # ── token + cost report (this run only) ──
    n = max(done, 1)
    pin, pcached, pout = PRICES.get(model, (None, None, None))
    uncached = tin - tcached
    print(f"\n=== TOKENS / COST ({done} calls, {wall:.0f}s) | model={model} ===")
    print(f"  input {tin:,} (cached {tcached:,} = {100*tcached/max(tin,1):.0f}%, uncached {uncached:,}) | output {tout:,}")
    print(f"  avg/question: in {tin/n:.0f} (cached {tcached/n:.0f}) | out {tout/n:.0f}")
    if pin is not None:
        cost = (uncached*pin + tcached*pcached + tout*pout) / 1e6
        per_q = cost / n
        print(f"  cost this run: ${cost:.4f}  (${per_q*1000:.4f}/1k questions)")
        print(f"  est. full 8,323: ${per_q*8323:.2f}")
        # also show a no-cache worst case for reference
        nocache = (tin*pin + tout*pout) / 1e6
        print(f"  (pricing {pin}/{pcached}/{pout} per 1M in/cached/out; no-cache would be ${nocache:.4f})")
    else:
        print(f"  (no PRICES entry for '{model}' — add one to report cost)")


if __name__ == "__main__":
    main()
