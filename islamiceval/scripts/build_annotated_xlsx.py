#!/usr/bin/env python3
"""
build_annotated_xlsx.py
=======================
Turn typology annotation files (annotate_typology.py output) into per-model Excel
files that show each answer next to its detected citation spans — so a human can
read the answer and its Quran/Hadith citations side by side.

For every  <model>__<annotator>_typology.json  in the annotations dir, writes
  <outdir>/<model>_sample_annotated.xlsx
with columns:  id | prompt | answer | model | citations

The `citations` cell lists one line per span (ordered by position):
    Quran/Correct  «وَابْتَغِ فِيمَا آتَاكَ ... الْمُفْسِدِينَ»  [225:436]
    Reference      «سورة البقرة 195»  [199:222]
The model's Correct/Incorrect verdict is shown verbatim (never overridden); a
span whose words could not be located verbatim is marked  ⚠.

Usage:
    python build_annotated_xlsx.py                       # annotator gpt-5.4, samples dir
    python build_annotated_xlsx.py --annotator gpt-5.4-mini
"""
import os, sys, glob, json, argparse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from verify_typology import load_corpora, norm, quran_exact, hadith_exact

HDR = ["id", "prompt", "answer", "model", "citations"]
WIDTHS = {"A": 10.0, "B": 50.0, "C": 90.0, "D": 16.0, "E": 90.0}


def fmt_span(sp):
    t = sp.get("type", "?")
    label = sp.get("label")
    # Quran/Hadith_Matn/Reference carry a Correct/Incorrect verdict; only Isnad = NA
    head = f"{t}/{label}" if label in ("Correct", "Incorrect") else t
    txt = sp.get("span_text", "")
    s, e = sp.get("span_start"), sp.get("span_end")
    return f"{head}  «{txt}»  [{s}:{e}]"


def citations_cell(rec):
    if rec.get("error"):
        return f"(annotation failed: {rec['error']})"
    spans = rec.get("spans") or []
    if not spans:
        return "(no citations)"
    spans = sorted(spans, key=lambda sp: (sp.get("span_start") if isinstance(sp.get("span_start"), int) else 0))
    return "\n".join(f"{i}. {fmt_span(sp)}" for i, sp in enumerate(spans, 1))


def apply_corpus_verdict(rec, corp):
    """Overwrite each Quran/Hadith_Matn span's Correct/Incorrect with the corpus
    verdict (silently — no marker). Reference/Isnad untouched. `corp` = load_corpora()."""
    QSET, QCONCAT, HSET, HCONCAT = corp
    ans = rec.get("answer", "") or ""
    for sp in (rec.get("spans") or []):
        t = sp.get("type")
        s, e = sp.get("span_start"), sp.get("span_end")
        if not (isinstance(s, int) and isinstance(e, int)):
            continue
        nslice = norm(ans[s:e])
        if t == "Quran":
            sp["label"] = "Correct" if quran_exact(nslice, QSET, QCONCAT) else "Incorrect"
        elif t == "Hadith_Matn":
            sp["label"] = "Correct" if hadith_exact(nslice, HSET, HCONCAT) else "Incorrect"


TYPES = ["Quran", "Hadith_Matn", "Hadith_Isnad", "Reference"]

def count_labels(ps):
    """counts[type][label] over all spans. label in {Correct, Incorrect, NA, other}."""
    counts = {t: {"Correct": 0, "Incorrect": 0, "NA": 0} for t in TYPES}
    for rec in ps:
        for sp in (rec.get("spans") or []):
            t = sp.get("type"); lab = sp.get("label")
            if t in counts and lab in counts[t]:
                counts[t][lab] += 1
            elif t in counts:
                counts[t].setdefault(lab, 0); counts[t][lab] += 1
    return counts


def build_one(typ_path, outdir, corp=None):
    data = json.load(open(typ_path, encoding="utf-8"))
    ps = data.get("per_sample", [])
    if corp is not None:
        for rec in ps:
            apply_corpus_verdict(rec, corp)
    model = data.get("answer_model") or "model"
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    ws.append(HDR)
    for c in ws[1]:
        c.font = Font(bold=True); c.alignment = Alignment(horizontal="center")
    top = Alignment(vertical="top")
    top_wrap = Alignment(vertical="top", wrap_text=True)
    for rec in ps:
        ws.append([rec.get("sample_id"), rec.get("prompt"), rec.get("answer"),
                   model, citations_cell(rec)])
    for col, w in WIDTHS.items():
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2):
        row[0].alignment = top                       # id
        row[1].alignment = top_wrap                  # prompt
        row[2].alignment = top_wrap                  # answer
        row[3].alignment = top                       # model
        row[4].alignment = top_wrap                  # citations

    # ── correctness summary at the bottom ──
    counts = count_labels(ps)
    ws.append([]); ws.append([])
    ws.append([f"SUMMARY — {model}", "Correct", "Incorrect", "NA", "total"])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    for t in TYPES:
        co, ic, na = counts[t]["Correct"], counts[t]["Incorrect"], counts[t]["NA"]
        ws.append([t, co, ic, na, co + ic + na])
    tot = {k: sum(counts[t][k] for t in TYPES) for k in ("Correct", "Incorrect", "NA")}
    ws.append(["ALL", tot["Correct"], tot["Incorrect"], tot["NA"],
               sum(tot.values())])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)

    out = os.path.join(outdir, f"{model}_sample_annotated.xlsx")
    wb.save(out)
    n_sp = sum(len(r.get("spans") or []) for r in ps)
    return model, len(ps), n_sp, out, counts


def main():
    ap = argparse.ArgumentParser(description="Build per-model annotated Excel from typology JSONs")
    ap.add_argument("--annotator", default="gpt-5.4")
    ap.add_argument("--anndir", default="../outputs/annotations")
    ap.add_argument("--outdir", default="../outputs/answers/samples")
    ap.add_argument("--no-correct", action="store_true",
                    help="keep the model's Correct/Incorrect instead of overwriting with the corpus verdict")
    args = ap.parse_args()

    pat = os.path.join(args.anndir, f"*__{args.annotator}_typology.json")
    files = sorted(glob.glob(pat))
    if not files:
        raise SystemExit(f"No typology files match {pat}")
    os.makedirs(args.outdir, exist_ok=True)
    corp = None if args.no_correct else load_corpora()
    mode = "model labels" if args.no_correct else "corpus-corrected Quran+Hadith verdicts"
    print(f"annotator={args.annotator} | {len(files)} files | {mode}\n")
    all_counts = {}
    for f in files:
        model, n, n_sp, out, counts = build_one(f, args.outdir, corp)
        all_counts[model] = counts
        print(f"  {model:24} {n:>3} answers, {n_sp:>4} spans -> {os.path.basename(out)}")

    # ── combined per-model, per-span-type Correct/Incorrect table ──
    print(f"\n=== CORRECT / INCORRECT per model per span type ({mode}) ===")
    print(f"{'model':24}" + "".join(f"{t[:9]:>20}" for t in TYPES))
    print(f"{'':24}" + "".join(f"{'C / I / NA':>20}" for _ in TYPES))
    print("-" * (24 + 20 * len(TYPES)))
    for model in sorted(all_counts):
        c = all_counts[model]
        cells = ""
        for t in TYPES:
            cell = f"{c[t]['Correct']}/{c[t]['Incorrect']}/{c[t]['NA']}"
            cells += f"{cell:>20}"
        print(f"{model:24}{cells}")


if __name__ == "__main__":
    main()
