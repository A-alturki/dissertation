#!/usr/bin/env python3
"""Consolidate the correction human-eval into ONE wide table.

Joins the blind item set (`correction_annotation.json` — question, answer, span, and each
model's proposed correction) with every annotator export (JSON or CSV) found in
outputs/correction/, producing one row per span with:

    id | model | ref | error_type | question | answer | span
    | gpt_correction | gemini_correction
    | <annotator> — GPT verdict | <annotator> — GPT correction
    | <annotator> — Gemini verdict | <annotator> — Gemini correction        (repeated per annotator)

first corrector = GPT-5.4, second = Gemini-3.1-pro. A model that declined a span
(verdict "خطأ", no proposed text) shows "خطأ — لا تصحيح مقترح" in its correction column.
Annotator verdict is correct / wrong / "" (unjudged); their manual fix (when they marked
"wrong") goes in the adjacent correction column.

Usage:
  python consolidate_correction_annotations.py                 # -> ...consolidated.xlsx
  python consolidate_correction_annotations.py --csv           # also write a .csv
  python consolidate_correction_annotations.py --out PATH
"""
import argparse
import csv
import glob
import json
import os
import sys

sys.stdout.reconfigure(encoding="utf-8")

HERE = os.path.dirname(os.path.abspath(__file__))
CORR = os.path.join(HERE, "..", "outputs", "correction")
JOIN = " ||| "  # separator for multiple corrections in one cell


# ---------------------------------------------------------------- items
def fmt_model_corrections(corrector):
    """corrector = {verdict, corrections:[{text, sources}]} -> display string."""
    if not corrector:
        return ""
    corr = corrector.get("corrections") or []
    if corr:
        parts = []
        for c in corr:
            txt = (c.get("text") or "").strip()
            src = c.get("sources") or []
            parts.append(f"{txt}  [{'، '.join(src)}]" if src else txt)
        return JOIN.join(parts)
    # no proposed text
    if corrector.get("verdict") == "خطأ":
        return "خطأ — لا تصحيح مقترح"
    return ""


def load_items(path):
    d = json.load(open(path, encoding="utf-8"))
    items = d["items"] if isinstance(d, dict) else d
    rows = {}
    for it in items:
        key = (it["id"], it["model"], (it["span"] or "").strip())
        cor = it.get("correctors", {})
        rows[key] = {
            "id": it["id"],
            "model": it["model"],
            "ref": it.get("ref", ""),
            "error_type": it.get("error_type", "") or "",
            "question": it.get("question", ""),
            "answer": it.get("answer", ""),
            "span": it.get("span", ""),
            "gpt_correction": fmt_model_corrections(cor.get("first")),      # first = GPT-5.4
            "gemini_correction": fmt_model_corrections(cor.get("second")),  # second = Gemini
        }
    return rows


# ---------------------------------------------------------------- annotators
def norm_verdict(v):
    v = (str(v).strip().lower() if v is not None else "")
    if v == "correct":
        return "صحيح"
    if v == "wrong":
        return "خطأ"
    return ""  # unjudged


def load_annotator(path):
    """-> (name, {(id,model,span,corrector): {'verdict','manual'}})"""
    recs = {}
    name = os.path.splitext(os.path.basename(path))[0].replace("correction_annotations", "").strip(" _")
    if path.lower().endswith(".json"):
        d = json.load(open(path, encoding="utf-8"))
        name = d.get("annotator", name)
        anns = d["annotations"] if isinstance(d, dict) else d
        for a in anns:
            k = (a["id"], a.get("model"), (a.get("span") or "").strip(), a.get("corrector"))
            man = [str(x).strip() for x in (a.get("manual_corrections") or []) if str(x).strip()]
            recs[k] = {"verdict": norm_verdict(a.get("judgment")), "manual": JOIN.join(man)}
    else:
        with open(path, encoding="utf-8-sig", newline="") as f:
            for r in csv.DictReader(f):
                if r.get("annotator"):
                    name = r["annotator"]
                k = (r["id"], r.get("model"), (r.get("span") or "").strip(), r.get("corrector"))
                man = (r.get("manual_corrections") or "").strip()
                recs[k] = {"verdict": norm_verdict(r.get("judgment")), "manual": man}
    return name, recs


def find_exports():
    # skip builder artifacts AND this script's own outputs
    skip = ("_first_topup", "_topup", "_original", "_combined", "_batch",
            "_consolidated", "_results")
    out = []
    for p in glob.glob(os.path.join(CORR, "correction_annotations*")):
        if (p.lower().endswith(".json") or p.lower().endswith(".csv")) \
                and not any(s in os.path.basename(p) for s in skip):
            out.append(p)
    return sorted(out)


# ---------------------------------------------------------------- build
def build(items, annotators):
    base_cols = ["id", "model", "ref", "error_type", "question", "answer", "span",
                 "gpt_correction", "gemini_correction"]
    anno_cols = []
    for name, _ in annotators:
        anno_cols += [f"{name} — GPT حكم", f"{name} — GPT تصحيح",
                      f"{name} — Gemini حكم", f"{name} — Gemini تصحيح"]
    header = base_cols + anno_cols

    rows = []
    for key, it in items.items():
        row = [it[c] for c in base_cols]
        id_, model, span = key
        for _, recs in annotators:
            for cor in ("first", "second"):   # first=GPT, second=Gemini
                rec = recs.get((id_, model, span, cor)) or {}
                row.append(rec.get("verdict", ""))
                row.append(rec.get("manual", ""))
        rows.append(row)
    return header, rows


def write_xlsx(path, header, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "consolidated"
    ws.sheet_view.rightToLeft = True
    ws.append(header)
    for r in rows:
        ws.append(r)
    # style header
    hf = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="374151")
    for c in ws[1]:
        c.font = hf
        c.fill = fill
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.freeze_panes = "A2"
    # widths
    widths = {"id": 10, "model": 20, "ref": 8, "error_type": 13, "question": 40,
              "answer": 60, "span": 40, "gpt_correction": 40, "gemini_correction": 40}
    from openpyxl.utils import get_column_letter
    for i, name in enumerate(header, 1):
        base = name.split(" — ")[0]
        w = widths.get(name, 26 if "—" in name else 22)
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(path)


def write_csv(path, header, rows):
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--items", default=os.path.join(CORR, "correction_annotation.json"))
    ap.add_argument("--out", default=os.path.join(CORR, "correction_annotations_consolidated.xlsx"))
    ap.add_argument("--csv", action="store_true", help="also write a .csv alongside the xlsx")
    args = ap.parse_args()

    items = load_items(args.items)
    exports = find_exports()
    annotators = [load_annotator(p) for p in exports]
    print(f"items: {len(items)} spans | annotators ({len(annotators)}): "
          + ", ".join(n for n, _ in annotators))

    header, rows = build(items, annotators)
    write_xlsx(args.out, header, rows)
    print(f"wrote {args.out}  ({len(rows)} rows x {len(header)} cols)")
    if args.csv:
        cpath = os.path.splitext(args.out)[0] + ".csv"
        write_csv(cpath, header, rows)
        print(f"wrote {cpath}")


if __name__ == "__main__":
    main()
