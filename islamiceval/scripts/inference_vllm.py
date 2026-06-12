"""
Stage 2: LLM Answer Generation (vLLM)
High-throughput batch inference for large-scale answer generation on the cluster.
Preferred over inference.py when running all 658+ prompts against multiple models.

Usage:
    python inference_vllm.py --model qwen3-8b
    python inference_vllm.py --model llama-3.3-70b --tensor-parallel 4
"""

import os, json, argparse
# Turing (sm_75): also disable the FlashInfer top-k/top-p sampler — like the
# attention backend, it imports flashinfer (uninstalled / can't JIT here). Must be
# set before vLLM is imported so the EngineCore subprocess inherits it.
os.environ.setdefault("VLLM_USE_FLASHINFER_SAMPLER", "0")
from vllm import LLM, SamplingParams
from vllm.v1.attention.backends.registry import AttentionBackendEnum

NO_SYSTEM_ROLE = {"jais-13b", "jais-70b", "acegpt-8b", "acegpt-32b", "acegpt-70b"}

# Some tokenizers ship no chat_template — supply the model's prompt format
# manually (built directly, not via apply_chat_template).
#   jais: its documented [|Human|]/[|AI|] Instruction format.
#   acegpt-v2: standard Llama-3 instruct format — it's Llama-3-8B based and its
#              tokenizer just omits the template (matches how llama-3.1-8b is fed).
JAIS_PROMPT_TEMPLATE = (
    "### Instruction: {system}\n\n"
    "أكمل المحادثة أدناه بين [|Human|] و [|AI|]:\n"
    "### Input: [|Human|] {prompt}\n### Response: [|AI|]"
)
LLAMA3_PROMPT_TEMPLATE = (
    "<|begin_of_text|><|start_header_id|>system<|end_header_id|>\n\n"
    "{system}<|eot_id|><|start_header_id|>user<|end_header_id|>\n\n"
    "{prompt}<|eot_id|><|start_header_id|>assistant<|end_header_id|>\n\n"
)
MANUAL_TEMPLATES = {
    "jais-13b":             JAIS_PROMPT_TEMPLATE,
    "jais-70b":             JAIS_PROMPT_TEMPLATE,
    "acegpt-8b":            LLAMA3_PROMPT_TEMPLATE,
    "acegpt-32b":           LLAMA3_PROMPT_TEMPLATE,
    "acegpt-70b":           LLAMA3_PROMPT_TEMPLATE,
    "deepseek-r1-llama-8b": LLAMA3_PROMPT_TEMPLATE,
}

# Some checkpoints ship a tokenizer that won't tokenize Arabic under transformers
# v5 (encodes to 0 tokens). deepseek-r1-distill-llama-8b is Llama-3.1-8B based and
# reuses Llama-3's vocab + EOS (128001), so use Llama-3.1's known-good tokenizer.
TOKENIZER_OVERRIDE = {
    "deepseek-r1-llama-8b": "meta-llama/Llama-3.1-8B-Instruct",
}

THINKING_KWARGS = {
    "fanar-2-27b":           {"no_thinking": True},
    "qwen3-0.6b":            {"enable_thinking": False},
    "qwen3-1.7b":            {"enable_thinking": False},
    "qwen3-4b":              {"enable_thinking": False},
    "qwen3-8b":              {"enable_thinking": False},
    "qwen3-14b":             {"enable_thinking": False},
    "qwen3-30b-a3b":         {"enable_thinking": False},
    "qwen3-32b":             {"enable_thinking": False},
    "qwen3.5-0.8b":          {"enable_thinking": False},
    "qwen3.5-2b":            {"enable_thinking": False},
    "qwen3.5-4b":            {"enable_thinking": False},
    "qwen3.5-9b":            {"enable_thinking": False},
    "qwen3.5-27b":           {"enable_thinking": False},
    "qwen3.5-35b-a3b":       {"enable_thinking": False},
    # Gemma 4 has a built-in thinking mode; disable it for clean citation answers
    # (same chat_template kwarg as Qwen3).
    "gemma-4-e4b":           {"enable_thinking": False},
    "gemma-4-12b":           {"enable_thinking": False},
    "gemma-4-26b-a4b":       {"enable_thinking": False},
    "gemma-4-31b":           {"enable_thinking": False},
}

STRIP_THINKING = {"deepseek-r1-llama-8b", "deepseek-r1-qwen-32b", "deepseek-r1-llama-70b",
                  "lfm2.5-8b-a1b"}  # LFM2.5 always emits CoT; verify its delimiter is <think>…</think>

# Models whose mandatory chain-of-thought needs a big output budget — 512 isn't enough
# for the model to finish thinking AND answer (it gets cut off mid-<think>). The context
# window (resolve_max_model_len cap) is auto-expanded to fit prompt + this many tokens.
MAX_TOKENS_OVERRIDE = {
    "lfm2.5-8b-a1b": 8192,   # LFM2.5's own examples use max_new_tokens=8192
}

# Gemma-3 is multimodal (vision+text). vLLM profiles the vision encoder at
# startup even for text-only inference, which OOMs on small MIG slices.
# Passing limit_mm_per_prompt={"image": 0} skips that profiling.
# All other models in MODELS are text-only — don't pass this param to them.
MULTIMODAL_MODELS = {"gemma-3-4b", "gemma-3-12b", "gemma-3-27b",
                     "gemma-4-e4b", "gemma-4-12b", "gemma-4-26b-a4b", "gemma-4-31b",
                     "llama-4-scout", "llama-4-maverick"}

SYSTEM_PROMPT = (
    "أنت مساعد إسلامي متخصص. أجب على السؤال بشكل دقيق ومختصر، "
    "مستنداً إلى القرآن الكريم والأحاديث النبوية الشريفة.\n"
    "عند الاستشهاد بآية قرآنية، اذكر اسم السورة ورقم الآية.\n"
    "عند الاستشهاد بحديث، اذكر المصدر (البخاري، مسلم، إلخ) إن أمكن."
)

ALT_PROMPT_2025 = "أجب عن السؤال التالي و استشهد بآيات من القرآن الكريم و احاديث شريفة"

ALT_PROMPT_2025_explicit = "أجب عن السؤال التالي و استشهد بآيات من القرآن الكريم و احاديث شريفة عند الاستشهاد بآية قرآنية، اذكر اسم السورة ورقم الآية. عند الاستشهاد بحديث، اذكر المصدر (البخاري، مسلم، إلخ)"

# Selectable system prompts (choose via --prompt; non-default writes to <model>_<prompt>.json).
PROMPTS = {
    "default":          SYSTEM_PROMPT,
    "alt2025":          ALT_PROMPT_2025,
    "alt2025-explicit": ALT_PROMPT_2025_explicit,
}

# Sampling fallbacks for models whose generation_config doesn't enable sampling.
DEFAULT_TEMPERATURE = 0.7
DEFAULT_TOP_P       = 0.9


def resolve_sampling(model_id: str, cli_temperature, cli_top_p):
    """Return (temperature, top_p) to sample with.

    We always sample. If the model's own generation_config enables sampling
    (do_sample=True), use its temperature/top_p; otherwise fall back to the
    DEFAULT_* values. A CLI flag, if given, overrides either source.
    """
    temperature, top_p = cli_temperature, cli_top_p
    if temperature is None or top_p is None:
        try:
            from transformers import GenerationConfig
            gc = GenerationConfig.from_pretrained(model_id)
            samples = bool(getattr(gc, "do_sample", False))
        except Exception:
            gc, samples = None, False
        if temperature is None:
            temperature = gc.temperature if (samples and gc.temperature) else DEFAULT_TEMPERATURE
        if top_p is None:
            top_p = gc.top_p if (samples and gc.top_p) else DEFAULT_TOP_P
    return temperature, top_p


def resolve_max_model_len(model_id, cap=4096):
    """Cap context at `cap`, but never exceed what the model actually supports
    (e.g. jais-13b is 2048). vLLM errors if max_model_len > the model's limit."""
    try:
        from transformers import AutoConfig
        m = getattr(AutoConfig.from_pretrained(model_id, trust_remote_code=True),
                    "max_position_embeddings", None)
        return min(cap, m) if m else cap
    except Exception:
        return cap


def load_prompts(path):
    """Load prompts as a list of {'id', 'prompt'}.
    .xlsx -> reads the 'qid' and 'prompt' columns; .json -> list of {'id','prompt'}.
    """
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        from openpyxl import load_workbook
        rows = load_workbook(path, read_only=True, data_only=True).active.iter_rows(values_only=True)
        header = list(next(rows))
        qi, pi = header.index("qid"), header.index("prompt")
        return [{"id": r[qi], "prompt": r[pi]} for r in rows if r[pi]]
    if ext == ".json":
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    raise ValueError(f"Unsupported input '{ext}' (use .xlsx or .json)")


MODELS = {
    # ==================== Arabic-centric ====================
    "allam-7b":              "ALLaM-AI/ALLaM-7B-Instruct-preview",
    "yehia-7b":              "Navid-AI/Yehia-7B-preview",        # ALLaM-7B based (Llama arch)
    "jais-13b":              "inceptionai/jais-13b-chat",
    "acegpt-8b":             "FreedomIntelligence/AceGPT-v2-8B-Chat",
    "acegpt-32b":            "FreedomIntelligence/AceGPT-v2-32B-Chat",   # v2; tp=2 on RTX 8000
    "acegpt-70b":            "FreedomIntelligence/AceGPT-v2-70B-Chat",   # v2; tp=4 (+maybe quant)
    "silma-9b":              "silma-ai/SILMA-9B-Instruct-v1.0",
    "fanar-1-9b":            "QCRI/Fanar-1-9B-Instruct",
    "fanar-2-27b":           "QCRI/Fanar-2-27B-Instruct",

    # ==================== Qwen family ======================
    "qwen3-0.6b":            "Qwen/Qwen3-0.6B",
    "qwen3-1.7b":            "Qwen/Qwen3-1.7B",
    "qwen3-4b":              "Qwen/Qwen3-4B",
    "qwen3-8b":              "Qwen/Qwen3-8B",
    "qwen3-14b":             "Qwen/Qwen3-14B",
    "qwen3-30b-a3b":         "Qwen/Qwen3-30B-A3B",           # MoE, 3B active

    # ==================== Qwen3.5 family (2026-02) =========
    "qwen3.5-0.8b":          "Qwen/Qwen3.5-0.8B",
    "qwen3.5-2b":            "Qwen/Qwen3.5-2B",
    "qwen3.5-4b":            "Qwen/Qwen3.5-4B",
    "qwen3.5-9b":            "Qwen/Qwen3.5-9B",
    "qwen3.5-27b":           "Qwen/Qwen3.5-27B",             # tp=2 on RTX 8000
    "qwen3.5-35b-a3b":       "Qwen/Qwen3.5-35B-A3B",         # MoE, 3B active; tp=2

    # ==================== Llama family (gated) ==============
    "llama-3.2-3b":          "meta-llama/Llama-3.2-3B-Instruct",
    "llama-3.1-8b":          "meta-llama/Llama-3.1-8B-Instruct",
    "llama-4-scout":         "meta-llama/Llama-4-Scout-17B-16E-Instruct",  # MoE

    # ==================== Gemma family (gated) ==============
    "gemma-3-4b":            "google/gemma-3-4b-it",
    "gemma-3-12b":           "google/gemma-3-12b-it",
    "gemma-3-27b":           "google/gemma-3-27b-it",
    # Gemma 4 (released 2026-04-02; multimodal, has a "thinking" mode). Supported by
    # vLLM >=0.19 + transformers >=5.5.0 — the box (vLLM 0.22 / tf 5.10) qualifies.
    # On Turing use --attention-backend FLEX_ATTENTION (head_dim=256 overflows Triton).
    "gemma-4-e4b":           "google/gemma-4-E4B-it",
    "gemma-4-12b":           "google/gemma-4-12B-it",
    "gemma-4-26b-a4b":       "google/gemma-4-26B-A4B-it",   # MoE, 4B active
    "gemma-4-31b":           "google/gemma-4-31B-it",

    # ==================== Mistral family ====================
    "mistral-7b":            "mistralai/Mistral-7B-Instruct-v0.3",
    "mistral-small-24b":     "mistralai/Mistral-Small-Instruct-2409",
    "mixtral-8x7b":          "mistralai/Mixtral-8x7B-Instruct-v0.1",  # MoE

    # ==================== DeepSeek family ===================
    "deepseek-r1-llama-8b":  "deepseek-ai/DeepSeek-R1-Distill-Llama-8B",

    # ==================== Other =============================
    "phi-3.5":             "microsoft/Phi-3.5-mini-instruct",
    "phi-4-14b":             "microsoft/phi-4",
    "glm-4-9b":              "THUDM/glm-4-9b-chat",
    "command-r-7b":          "CohereForAI/c4ai-command-r7b-12-2024",
    "command-r-7b-arabic":   "CohereForAI/c4ai-command-r7b-arabic-02-2025",
    "lfm2.5-8b-a1b":         "LiquidAI/LFM2.5-8B-A1B",       # hybrid conv+attn MoE, 1B active; always emits CoT

    # ==========================================================
    # HEAVY COMPUTE — need A100 80GB+ or multi-GPU
    # ==========================================================

    # 1x A100 80GB
    "qwen3-32b":             "Qwen/Qwen3-32B",
    "deepseek-r1-qwen-32b":  "deepseek-ai/DeepSeek-R1-Distill-Qwen-32B",
    "llama-3.3-70b":         "meta-llama/Llama-3.3-70B-Instruct",        # tight, may need quantization
    "jais-70b":              "inceptionai/jais-adapted-70b-chat",

    # 2x A100 80GB
    "deepseek-r1-llama-70b": "deepseek-ai/DeepSeek-R1-Distill-Llama-70B",

    # 4x A100 80GB
    "qwen3-235b":            "Qwen/Qwen3-235B-A22B",                     # MoE, 22B active
    "llama-4-maverick":      "meta-llama/Llama-4-Maverick-17B-128E-Instruct",  # MoE
    "deepseek-v3":           "deepseek-ai/DeepSeek-V3-0324",             # 685B MoE, 37B active
}

islamic_eval_models = ["allam-7b", "jais-13b", "acegpt-8b", "silma-9b", "fanar-1-9b", "qwen3-8b", "gemma-3-12b", "mistral-7b", "deepseek-r1-llama-8b", "llama-3.1-8b"]

### NEXT MODELS FOR COMPLETE ANALYSIS
# "fanar-2-27b", "qwen3-14b/27B", "qwen3.5-9B/27B" ,  "llama-4-scout", "gemma-3-27b", "mistral-small-24b", "mixtral-8x7b",
# "LFM2.5-8B-A1B", "Gemma 4 12B", "AceGPT V2 7B, 8B, 13B, 32B" Applied-Innovation-Center/AIC-1 33B", "CohereForAI/c4ai-command-r7b-arabic-02-2025", "Yehia-7B-preview"
# "Jais 30B", "CohereForAI/aya-expanse-32b",  

# Named groups you can pass to --model to run several models back-to-back.
MODEL_GROUPS = {
    "islamic-eval": islamic_eval_models,
    "all":          list(MODELS.keys()),
}


def run_group(args, model_names):
    """Run several models sequentially, each in its own fresh subprocess so every
    model gets a clean GPU (vLLM doesn't reliably free VRAM within one process)."""
    import subprocess, sys
    print(f"Running {len(model_names)} models: {model_names}")
    failed = []
    for i, m in enumerate(model_names, 1):
        cmd = [sys.executable, os.path.abspath(__file__),
               "--model", m,
               "--input", args.input,
               "--output-dir", args.output_dir,
               "--max-tokens", str(args.max_tokens),
               "--batch-size", str(args.batch_size),
               "--tensor-parallel", str(args.tensor_parallel),
               "--attention-backend", args.attention_backend]
        if args.temperature is not None: cmd += ["--temperature", str(args.temperature)]
        if args.top_p       is not None: cmd += ["--top-p",       str(args.top_p)]
        cmd += ["--prompt", args.prompt]
        if args.limit is not None: cmd += ["--limit", str(args.limit)]
        print(f"\n{'='*60}\n[{i}/{len(model_names)}] {m}\n{'='*60}")
        if subprocess.run(cmd).returncode != 0:
            print(f"[WARN] {m} failed — continuing with the rest.")
            failed.append(m)
    print(f"\nDone. {len(model_names) - len(failed)}/{len(model_names)} succeeded."
          + (f"  Failed: {failed}" if failed else ""))


def main():
    parser = argparse.ArgumentParser(description="Generate answers with vLLM (Stage 2)")
    parser.add_argument("--model",           required=True,
                        help="a model key, or a group: " + " | ".join(MODEL_GROUPS))
    parser.add_argument("--input",           default="../data/classified/rag_questions.json",
                        help="prompts file: .xlsx (qid+prompt) or .json")
    parser.add_argument("--limit",           type=int, default=None,
                        help="Use only the first N prompts (for quick experiments)")
    parser.add_argument("--output-dir",      default="../outputs/answers/explicit/")
    parser.add_argument("--max-tokens",      type=int, default=512)
    parser.add_argument("--batch-size",      type=int, default=128,
                        help="Checkpoint interval: generate and save this many prompts at a time")
    parser.add_argument("--temperature",     type=float, default=None,
                        help="Override sampling temperature (default: the model's own; pass 0 for greedy)")
    parser.add_argument("--top-p",           type=float, default=None,
                        help="Override nucleus sampling top-p (default: the model's own)")
    parser.add_argument("--tensor-parallel", type=int, default=1,
                        help="Number of GPUs for tensor parallelism (for 70B+ models)")
    parser.add_argument("--attention-backend", default="TRITON_ATTN",
                        help="vLLM attention backend (AttentionBackendEnum name). Default "
                             "TRITON_ATTN. Gemma (head_dim=256) needs FLEX_ATTENTION or TORCH_SDPA on Turing.")
    parser.add_argument("--enforce-eager", action="store_true",
                        help="Disable torch.compile/CUDA graphs. Use if engine init fails with a "
                             "Dynamo 'graph break' (e.g. Gemma 4 on Turing/sm_75).")
    parser.add_argument("--prompt", choices=list(PROMPTS), default="default",
                        help="System prompt to use. Non-default writes to <model>_<prompt>.json")
    args = parser.parse_args()

    prompts = load_prompts(args.input)
    if args.limit:
        prompts = prompts[:args.limit]
    print(f"Loaded {len(prompts)} prompts from {args.input}")

    model_id = MODELS[args.model]
    print(f"Loading {args.model} with vLLM ({model_id})  tp={args.tensor_parallel}")
    temperature, top_p = resolve_sampling(model_id, args.temperature, args.top_p)
    print(f"Sampling on — temperature={temperature}, top_p={top_p}")

    # Per-model output budget: CoT models (e.g. LFM2.5) need far more than the 512 default
    # or they never finish thinking. Expand the context window to fit prompt + that budget.
    eff_max_tokens = MAX_TOKENS_OVERRIDE.get(args.model, args.max_tokens)
    cap = (eff_max_tokens + 4096) if args.model in MAX_TOKENS_OVERRIDE else 4096
    max_len = resolve_max_model_len(model_id, cap=cap)
    llm_kwargs = dict(
        model=model_id,
        tensor_parallel_size=args.tensor_parallel,
        trust_remote_code=True,
        dtype="auto",
        max_model_len=max_len,
        gpu_memory_utilization=0.70,
        # Turing (RTX 8000, sm_75): avoid FlashInfer (uninstalled; can't JIT here) and
        # FLASH_ATTN (needs sm_80+). Default TRITON_ATTN; Gemma (head_dim=256) overflows
        # Triton's shared mem, so it needs FLEX_ATTENTION/TORCH_SDPA via --attention-backend.
        attention_backend=AttentionBackendEnum[args.attention_backend],
    )
    if args.enforce_eager:
        # Disable torch.compile / CUDA graphs — avoids Dynamo "graph break" failures at
        # engine init (e.g. Gemma 4 on Turing/sm_75). Slower, but robust.
        llm_kwargs["enforce_eager"] = True
    if args.model in MULTIMODAL_MODELS:
        llm_kwargs["limit_mm_per_prompt"] = {"image": 0}
    if args.model in TOKENIZER_OVERRIDE:
        llm_kwargs["tokenizer"] = TOKENIZER_OVERRIDE[args.model]
        print(f"Tokenizer override -> {llm_kwargs['tokenizer']}")

    llm = LLM(**llm_kwargs)
    tokenizer = llm.get_tokenizer()

    sampling = SamplingParams(temperature=temperature, top_p=top_p,
                              max_tokens=eff_max_tokens)

    # Prompt + output must fit in max_len; leave room for eff_max_tokens of output.
    # Over-long prompts (e.g. jais-13b's 2048 ctx) are left-truncated below,
    # keeping the tail (assistant cue / question end) and dropping the head.
    truncate_to = max(1, max_len - eff_max_tokens)

    # Active system prompt (selected via --prompt; see PROMPTS).
    system_prompt = PROMPTS[args.prompt]
    print(f"System prompt: {args.prompt}")

    def build_messages(item):
        if args.model in NO_SYSTEM_ROLE:
            return [{"role": "user", "content": f"{system_prompt}\n\n{item['prompt']}"}]
        return [
            {"role": "system", "content": system_prompt},
            {"role": "user",   "content": item["prompt"]},
        ]

    extra_template_kwargs = THINKING_KWARGS.get(args.model, {})

    conversations = []
    n_truncated = 0
    for item in prompts:
        try:
            manual_tmpl = MANUAL_TEMPLATES.get(args.model)
            if manual_tmpl:
                text = manual_tmpl.format(system=system_prompt, prompt=item["prompt"])
            else:
                text = tokenizer.apply_chat_template(
                    build_messages(item), tokenize=False, add_generation_prompt=True,
                    **extra_template_kwargs
                )
            # Enforce the context budget: left-truncate over-long prompts (keep the
            # tail/assistant cue) so prompt + max_tokens fits in max_len.
            ids = tokenizer(text, add_special_tokens=False).input_ids
            if len(ids) > truncate_to:
                text = tokenizer.decode(ids[-truncate_to:])
                n_truncated += 1
        except Exception as e:
            print(f"Template error on id={item['id']}: {e} — skipping")
            text = None
        conversations.append(text)
    if n_truncated:
        print(f"Left-truncated {n_truncated} over-long prompt(s) to {truncate_to} tokens.")

    valid_indices = [i for i, c in enumerate(conversations) if c is not None]
    valid_convs   = [conversations[i] for i in valid_indices]
    valid_prompts = [prompts[i] for i in valid_indices]
    skipped = len(prompts) - len(valid_indices)
    if skipped:
        print(f"Warning: {skipped} prompts skipped due to template errors.")

    import re
    def clean(text: str) -> str:
        if args.model in STRIP_THINKING:
            text = re.sub(r"<think>.*?</think>", "", text, flags=re.DOTALL)
        return text.strip()

    os.makedirs(args.output_dir, exist_ok=True)
    suffix = "" if args.prompt == "default" else f"_{args.prompt}"
    out_path = os.path.join(args.output_dir, f"{args.model}{suffix}.json")

    def save(results):
        # atomic: write to a temp file then replace, so a crash mid-write
        # can't corrupt the checkpoint.
        tmp = out_path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        os.replace(tmp, out_path)

    # Resume: keep answers already saved, regenerate only the missing ids.
    results = []
    if os.path.exists(out_path):
        with open(out_path, encoding="utf-8") as f:
            results = json.load(f)
        done = {r["id"] for r in results}
        keep = [(c, p) for c, p in zip(valid_convs, valid_prompts) if p["id"] not in done]
        valid_convs  = [c for c, _ in keep]
        valid_prompts = [p for _, p in keep]
        print(f"Resuming: {len(done)} already done, {len(valid_convs)} remaining.")

    total = len(valid_convs)
    print(f"Generating {total} answers in batches of {args.batch_size}...")
    for start in range(0, total, args.batch_size):
        convs_b   = valid_convs[start:start + args.batch_size]
        prompts_b = valid_prompts[start:start + args.batch_size]
        outputs = llm.generate(convs_b, sampling)
        results.extend(
            {
                "id":     item["id"],
                "prompt": item["prompt"],
                "answer": clean(out.outputs[0].text),
                "model":  args.model,
            }
            for item, out in zip(prompts_b, outputs)
        )
        save(results)
        print(f"  checkpoint: {min(start + args.batch_size, total)}/{total} done "
              f"-> {out_path} ({len(results)} total)")

    print(f"Saved {len(results)} answers -> {out_path}")


if __name__ == "__main__":
    main()
