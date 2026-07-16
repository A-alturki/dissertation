"""
Stage 2: LLM Answer Generation (HuggingFace Transformers)
Generates answers with Quran/Hadith citations for Islamic prompts.
Uses 4-bit quantization (AWQ/NF4) to fit 7-13B models on 2080 Ti (11GB VRAM).

Usage:
    python inference_hugging_face.py --model allam-7b --input ../data/classified/rag_questions.json
    python inference_hugging_face.py --model qwen3-8b --no-quantize   # for A6000 (48GB)
"""

import os, sys, json, argparse
# Reduce CUDA fragmentation (helps avoid OOM on variable-length batches). Must be
# set before torch initializes CUDA.
os.environ.setdefault("PYTORCH_CUDA_ALLOC_CONF", "expandable_segments:True")
import torch
from transformers import AutoTokenizer, AutoModelForCausalLM, BitsAndBytesConfig
from tqdm.auto import tqdm

# --- compat shim: transformers v5 removed find_pruneable_heads_and_indices from
# pytorch_utils, but the jais-family remote modeling code still imports it at load
# time (ImportError kills jais-13b/30b on both sdpa+eager). Restore the (stable) fn. ---
import transformers.pytorch_utils as _tpu
if not hasattr(_tpu, "find_pruneable_heads_and_indices"):
    def find_pruneable_heads_and_indices(heads, n_heads, head_size, already_pruned_heads):
        mask = torch.ones(n_heads, head_size)
        heads = set(heads) - already_pruned_heads
        for head in heads:
            head = head - sum(1 if h < head else 0 for h in already_pruned_heads)
            mask[head] = 0
        mask = mask.view(-1).contiguous().eq(1)
        index = torch.arange(len(mask))[mask].long()
        return heads, index
    _tpu.find_pruneable_heads_and_indices = find_pruneable_heads_and_indices

# Models whose chat templates don't accept a system role — system prompt is
# folded into the first user message instead.
NO_SYSTEM_ROLE = {"jais-13b", "jais-70b", "acegpt-8b"}

# Per-model extra kwargs passed to apply_chat_template to disable thinking output.
# DeepSeek-R1 models have no template flag — <think> blocks are stripped in post-processing.
THINKING_KWARGS = {
    "fanar-2-27b":           {"no_thinking": True},
    "qwen3-0.6b":            {"enable_thinking": False},
    "qwen3-1.7b":            {"enable_thinking": False},
    "qwen3-4b":              {"enable_thinking": False},
    "qwen3-8b":              {"enable_thinking": False},
    "qwen3-14b":             {"enable_thinking": False},
    "qwen3-30b-a3b":         {"enable_thinking": False},
    "qwen3-32b":             {"enable_thinking": False},
    # Gemma 4 has a built-in thinking mode — disable for clean citation answers.
    "gemma-4-e4b":           {"enable_thinking": False},
    "gemma-4-12b":           {"enable_thinking": False},
    "gemma-4-26b-a4b":       {"enable_thinking": False},
    "gemma-4-31b":           {"enable_thinking": False},
}

# Models whose <think> blocks must be stripped in post-processing (no template flag).
STRIP_THINKING = {"deepseek-r1-llama-8b", "deepseek-r1-qwen-32b", "deepseek-r1-llama-70b"}

# Some tokenizers ship no chat_template — supply the model's prompt format manually
# (built directly, not via apply_chat_template). Ported from inference_vllm.py.
#   jais: its documented [|Human|]/[|AI|] Instruction format.
#   acegpt-v2: standard Llama-3 instruct format (it's Llama-3-8B based; tokenizer omits the template).
JAIS_PROMPT_TEMPLATE = (
    "### Instruction: {system}\n\n"
    "أكمل المحادثة أدناه بين [|Human|] و [|AI|]:\n"
    "### Input: [|Human|] {prompt}\n### Response: [|AI|]"
)
LLAMA3_PROMPT_TEMPLATE = (
    "<|begin_of_text|><|start_header_id|>system<|end_header_id|>\n\n"
    "{system}<|eot_id|><|start_header_id|>user<|end_header_id|>\n\n"
    "{prompt}<|eot_id|><|start_header_id|>assistant<|end_header_id|>\n\n"
)
MANUAL_TEMPLATES = {
    "jais-13b":  JAIS_PROMPT_TEMPLATE,
    "jais-70b":  JAIS_PROMPT_TEMPLATE,
    "jais-30b":  JAIS_PROMPT_TEMPLATE,   # jais-family-30b ships no chat_template (verified)
    "acegpt-8b": LLAMA3_PROMPT_TEMPLATE,
    # NB: acegpt-32b DOES ship a chat_template (unlike the 8B) -> uses apply_chat_template.
}

# Set by --allow-thinking: when True, do NOT inject the thinking-disabling kwargs, so
# models with a native thinking mode (e.g. Qwen3) reason before answering.
ALLOW_THINKING = False

# Resolved at runtime from each model's creator-default generation_config (resolve_gen_config).
GEN_TOP_K = None
GEN_DO_SAMPLE = None

# Model-card sampling recommendations for fields the repo's generation_config.json omits.
CARD_OVERRIDES = {
    "ministral-3-14b": {"temperature": 0.15},   # Mistral model-card recommendation
}

SYSTEM_PROMPT = (
    "أنت مساعد إسلامي متخصص. أجب على السؤال بشكل دقيق ومختصر، "
    "مستنداً إلى القرآن الكريم والأحاديث النبوية الشريفة.\n"
    "عند الاستشهاد بآية قرآنية، اذكر اسم السورة ورقم الآية.\n"
    "عند الاستشهاد بحديث، اذكر المصدر (البخاري، مسلم، إلخ) إن أمكن."
)

ALT_PROMPT_2025 = "أجب عن السؤال التالي و استشهد بآيات من القرآن الكريم و احاديث شريفة"

ALT_PROMPT_2025_explicit = "أجب عن السؤال التالي و استشهد بآيات من القرآن الكريم و احاديث شريفة عند الاستشهاد بآية قرآنية، اذكر اسم السورة ورقم الآية. عند الاستشهاد بحديث، اذكر المصدر (البخاري، مسلم، إلخ)"

VANILLA_PROMPT =  "أجب عن السؤال التالي."



PROMPTS = {"default": SYSTEM_PROMPT, "alt2025": ALT_PROMPT_2025,
           "alt2025-explicit": ALT_PROMPT_2025_explicit, "vanilla": VANILLA_PROMPT}

# Sampling fallbacks for models whose generation_config doesn't enable sampling.
DEFAULT_TEMPERATURE = 0.7
DEFAULT_TOP_P       = 0.9


def resolve_sampling(model_id: str, cli_temperature, cli_top_p):
    """Return (temperature, top_p) to sample with.

    We always sample. If the model's own generation_config enables sampling
    (do_sample=True), use its temperature/top_p; otherwise fall back to the
    DEFAULT_* values. A CLI flag, if given, overrides either source.
    """
    temperature, top_p = cli_temperature, cli_top_p
    if temperature is None or top_p is None:
        try:
            from transformers import GenerationConfig
            gc = GenerationConfig.from_pretrained(model_id)
            samples = bool(getattr(gc, "do_sample", False))
        except Exception:
            gc, samples = None, False
        if temperature is None:
            temperature = gc.temperature if (samples and gc.temperature) else DEFAULT_TEMPERATURE
        if top_p is None:
            top_p = gc.top_p if (samples and gc.top_p) else DEFAULT_TOP_P
    return temperature, top_p


def resolve_gen_config(model_id, model_key, cli_temperature, cli_top_p):
    """Creator-default sampling: read the repo's generation_config.json, then apply the
    model card's recommendation for anything it omits (CARD_OVERRIDES), then a mild
    fallback — never impose greedy. CLI --temperature/--top-p override everything.
    Returns dict(do_sample, temperature, top_p, top_k)."""
    gc = None
    try:
        from transformers import GenerationConfig
        gc = GenerationConfig.from_pretrained(model_id)
    except Exception:
        pass
    do_sample   = bool(getattr(gc, "do_sample", False)) if gc else False
    temperature = getattr(gc, "temperature", None)      if gc else None
    top_p       = getattr(gc, "top_p", None)            if gc else None
    top_k       = getattr(gc, "top_k", None)            if gc else None
    for k, v in CARD_OVERRIDES.get(model_key, {}).items():   # card fills omitted fields
        if   k == "do_sample":            do_sample = do_sample or v
        elif k == "temperature" and not temperature: temperature = v; do_sample = True
        elif k == "top_p"       and not top_p:       top_p = v
        elif k == "top_k"       and not top_k:       top_k = v
    if not do_sample:                                   # never greedy -> mild sampling
        do_sample = True
        temperature = temperature or DEFAULT_TEMPERATURE
        top_p       = top_p or DEFAULT_TOP_P
    if do_sample and not temperature: temperature = DEFAULT_TEMPERATURE
    if do_sample and not top_p:       top_p = DEFAULT_TOP_P
    if cli_temperature is not None:                     # explicit CLI wins
        temperature = cli_temperature; do_sample = cli_temperature > 0
    if cli_top_p is not None: top_p = cli_top_p
    return {"do_sample": do_sample, "temperature": temperature, "top_p": top_p, "top_k": top_k}


def resolve_dtype(model_id, mode):
    """mode: fp16 | bf16 | native. 'native' reads the repo config's torch_dtype
    (bf16-native models stay bf16 — avoids Gemma fp16 soft-cap overflow)."""
    if mode == "fp16": return torch.float16
    if mode == "bf16": return torch.bfloat16
    try:
        from transformers import AutoConfig
        dt = str(getattr(AutoConfig.from_pretrained(model_id, trust_remote_code=True), "torch_dtype", ""))
        if "bfloat16" in dt: return torch.bfloat16
        if "float16"  in dt: return torch.float16
    except Exception:
        pass
    return torch.bfloat16   # 2026 large instruct models are bf16-native by default


_AR = None
def is_degenerate(a: str) -> bool:
    """Smoke-test guard: flag empty/NaN-ish/degenerate fp16 output."""
    import re
    global _AR
    if _AR is None: _AR = re.compile(r'[؀-ۿ]')
    a = (a or "").strip()
    if not a or a.startswith("ERROR") or len(a) < 5:      return True
    letters = [c for c in a if c.isalpha()]
    if letters and sum(1 for c in letters if _AR.match(c))/len(letters) < 0.3: return True  # not Arabic
    toks = a.split()
    if len(toks) >= 8 and len(set(toks)) <= 2:            return True   # single-token loop
    return False

# used same models as islamiceval + mistral/acegpt/silma 
MODELS = {
    # ==================== Arabic-centric ====================
    "allam-7b":              "ALLaM-AI/ALLaM-7B-Instruct-preview",
    "jais-13b":              "inceptionai/jais-13b-chat",
    "acegpt-8b":             "FreedomIntelligence/AceGPT-v2-8B-Chat",
    "silma-9b":              "silma-ai/SILMA-9B-Instruct-v1.0",
    "fanar-1-9b":            "QCRI/Fanar-1-9B-Instruct",
    "fanar-2-27b":           "QCRI/Fanar-2-27B-Instruct",

    # ==================== Qwen family ======================
    "qwen3-0.6b":            "Qwen/Qwen3-0.6B",
    "qwen3-1.7b":            "Qwen/Qwen3-1.7B",
    "qwen3-4b":              "Qwen/Qwen3-4B",
    "qwen3-8b":              "Qwen/Qwen3-8B",
    "qwen3-14b":             "Qwen/Qwen3-14B",
    "qwen3-30b-a3b":         "Qwen/Qwen3-30B-A3B",           # MoE, 3B active

    # ==================== Llama family (gated) ==============
    "llama-3.2-3b":          "meta-llama/Llama-3.2-3B-Instruct",
    "llama-3.1-8b":          "meta-llama/Llama-3.1-8B-Instruct",
    "llama-4-scout":         "meta-llama/Llama-4-Scout-17B-16E-Instruct",  # MoE

    # ==================== Gemma family (gated) ==============
    "gemma-3-4b":            "google/gemma-3-4b-it",
    "gemma-3-12b":           "google/gemma-3-12b-it",
    "gemma-3-27b":           "google/gemma-3-27b-it",
    # Gemma 4 (2026-04). HF path — proven route for Gemma on Turing when vLLM's
    # attention backends don't work (TRITON overflow / FLEX graph-break / no TORCH_SDPA).
    "gemma-4-e4b":           "google/gemma-4-E4B-it",
    "gemma-4-12b":           "google/gemma-4-12B-it",
    "gemma-4-26b-a4b":       "google/gemma-4-26B-A4B-it",
    "gemma-4-31b":           "google/gemma-4-31B-it",

    # ==================== Mistral family ====================
    "mistral-7b":            "mistralai/Mistral-7B-Instruct-v0.3",
    "mistral-small-24b":     "mistralai/Mistral-Small-Instruct-2409",
    "mixtral-8x7b":          "mistralai/Mixtral-8x7B-Instruct-v0.1",  # MoE

    # ==================== DeepSeek family ===================
    "deepseek-r1-llama-8b":  "deepseek-ai/DeepSeek-R1-Distill-Llama-8B",

    # ==================== Other =============================
    "phi-3.5":             "microsoft/Phi-3.5-mini-instruct",
    "phi-4-14b":             "microsoft/phi-4",
    "glm-4-9b":              "THUDM/glm-4-9b-chat",
    "command-r-7b":          "CohereForAI/c4ai-command-r7b-12-2024",
    

    # ==========================================================
    # HEAVY COMPUTE — need A100 80GB+ or multi-GPU
    # ==========================================================

    # 1x A100 80GB
    "qwen3-32b":             "Qwen/Qwen3-32B",
    "deepseek-r1-qwen-32b":  "deepseek-ai/DeepSeek-R1-Distill-Qwen-32B",
    "llama-3.3-70b":         "meta-llama/Llama-3.3-70B-Instruct",       
    "jais-70b":              "inceptionai/jais-adapted-70b-chat",

    # 2x A100 80GB
    "deepseek-r1-llama-70b": "deepseek-ai/DeepSeek-R1-Distill-Llama-70B",

    # 4x A100 80GB
    "qwen3-235b":            "Qwen/Qwen3-235B-A22B",                     # MoE, 22B active
    "llama-4-maverick":      "meta-llama/Llama-4-Maverick-17B-128E-Instruct",  # MoE
    "deepseek-v3":           "deepseek-ai/DeepSeek-V3-0324",             # 685B MoE, 37B active

    # ==== Larger 2026 additions — sequential HF run on 2×RTX 8000 (device_map=auto) ====
    "ministral-3-14b":       "mistralai/Ministral-3-14B-Instruct-2512-BF16",  # BF16 repo (NOT the FP8 default); multimodal->text path
    # "fanar-2-27b" already above (Gemma3ForCausalLM — bf16)
    # "gemma-4-31b"  already above -> "google/gemma-4-31B-it" (multimodal->text path; bf16)
    "qwen3.5-27b":           "Qwen/Qwen3.5-27B",                              # dense, multimodal->text path; thinking-capable
    "karnak-40b":            "Applied-Innovation-Center/Karnak-40B-v1.0",     # Qwen3-MoE arch
    "falcon-h1-34b":         "tiiuae/Falcon-H1-34B-Instruct",                 # hybrid (Mamba+attn); general (Arabic-34B repo doesn't exist)
    "acegpt-32b":            "FreedomIntelligence/AceGPT-v2-32B-Chat",        # Llama arch; HAS chat_template (unlike 8B)
    "jais-30b":              "inceptionai/jais-family-30b-16k-chat",          # gated; custom ALiBi; manual template
}


def load_model(model_id: str, quantize: bool = True, dtype=torch.float16, attn="auto"):
    bnb_config = None
    if quantize:
        bnb_config = BitsAndBytesConfig(
            load_in_4bit=True,
            bnb_4bit_compute_dtype=torch.float16,
            bnb_4bit_use_double_quant=True,
            bnb_4bit_quant_type="nf4",
        )
    tokenizer = AutoTokenizer.from_pretrained(model_id, trust_remote_code=True)
    # Batched decoder generation needs LEFT padding and a pad token.
    if tokenizer.pad_token_id is None:
        tokenizer.pad_token = tokenizer.eos_token
    tokenizer.padding_side = "left"
    # dtype: pass explicitly (else from_pretrained loads fp32 -> CPU-offload). device_map="auto"
    # shards across all VISIBLE GPUs (caller sets CUDA_VISIBLE_DEVICES). attn: "sdpa" is the
    # only fast attention on Turing/sm_75 (no FlashAttention); "auto" leaves it to transformers.
    def _load(cls, dkey):
        kw = dict(quantization_config=bnb_config, device_map="auto", trust_remote_code=True)
        if not quantize:
            kw[dkey] = dtype
        if attn and attn != "auto":
            kw["attn_implementation"] = attn
        return cls.from_pretrained(model_id, **kw)
    # Multimodal checkpoints ship a *ForConditionalGeneration config that AutoModelForCausalLM
    # rejects ("Unrecognized configuration") — fall back to the image-text-to-text auto class,
    # which generates text fine from text-only input_ids (no pixel inputs). Text models load first.
    classes = [AutoModelForCausalLM]
    try:
        from transformers import AutoModelForImageTextToText
        classes.append(AutoModelForImageTextToText)
    except Exception:
        pass
    last = None
    for cls in classes:
        try:
            try:
                return tokenizer, _load(cls, "dtype")       # transformers v5 arg name
            except TypeError:                               # older transformers: torch_dtype
                return tokenizer, _load(cls, "torch_dtype")
        except ValueError as e:
            last = e
            if "Unrecognized configuration" in str(e):
                print(f"  [load] {cls.__name__} rejected this config; trying the next auto-class…")
                continue
            raise
    raise last


def build_messages(prompt: str, model_key: str) -> list:
    if model_key in NO_SYSTEM_ROLE:
        # Fold system prompt into the user turn for models without system role support
        return [{"role": "user", "content": f"{SYSTEM_PROMPT}\n\n{prompt}"}]
    return [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user",   "content": prompt},
    ]


def generate_with_retry(prompts, tokenizer, model, model_key, max_new_tokens, temperature, top_p):
    """Run a batch; on CUDA OOM, free cache and recursively halve it so one long
    batch can't poison many answers. Only a lone prompt that still OOMs is recorded
    as an error."""
    try:
        return generate_batch(prompts, tokenizer, model, model_key, max_new_tokens, temperature, top_p)
    except torch.cuda.OutOfMemoryError:
        torch.cuda.empty_cache()
        if len(prompts) == 1:
            print(f"\n  [OOM] single prompt too large — recording ERROR")
            return ["ERROR: CUDA OOM on single prompt"]
        mid = len(prompts) // 2
        print(f"\n  [OOM] splitting batch of {len(prompts)} -> {mid}+{len(prompts)-mid}")
        return (generate_with_retry(prompts[:mid], tokenizer, model, model_key, max_new_tokens, temperature, top_p)
                + generate_with_retry(prompts[mid:], tokenizer, model, model_key, max_new_tokens, temperature, top_p))


def generate_batch(prompts, tokenizer, model, model_key: str, max_new_tokens: int = 512,
                   temperature: float = None, top_p: float = None) -> list:
    """Generate answers for a batch of prompts in one model.generate() call."""
    # --allow-thinking: skip the thinking-disabling kwargs so native-thinking models reason.
    extra = {} if ALLOW_THINKING else THINKING_KWARGS.get(model_key, {})
    manual = MANUAL_TEMPLATES.get(model_key)
    if manual:                                   # tokenizer has no chat_template
        texts = [manual.format(system=SYSTEM_PROMPT, prompt=p) for p in prompts]
    else:
        texts = [
            tokenizer.apply_chat_template(build_messages(p, model_key), tokenize=False,
                                          add_generation_prompt=True, **extra)
            for p in prompts
        ]
    # apply_chat_template output and the Llama-3 manual template already embed BOS/special
    # tokens (-> don't re-add); the jais manual template does not, so let the tokenizer add them.
    add_special = bool(manual) and not manual.startswith("<|begin_of_text|>")
    inputs = tokenizer(texts, return_tensors="pt", padding=True, truncation=True,
                       max_length=4096, add_special_tokens=add_special).to(model.device)

    # do_sample / top_k come from the model's resolved creator-default config (set in main);
    # fall back to temperature>0 when unset (original behaviour).
    do_sample = GEN_DO_SAMPLE if GEN_DO_SAMPLE is not None else (temperature > 0)
    with torch.no_grad():
        output = model.generate(
            **inputs,
            max_new_tokens=max_new_tokens,
            do_sample=do_sample,
            temperature=temperature if do_sample else None,
            top_p=top_p if do_sample else None,
            top_k=GEN_TOP_K if (do_sample and GEN_TOP_K) else None,
            pad_token_id=tokenizer.pad_token_id,
        )
    gen = output[:, inputs["input_ids"].shape[1]:]   # left-padded -> same input width for all
    answers = tokenizer.batch_decode(gen, skip_special_tokens=True)
    if model_key in STRIP_THINKING:
        import re
        answers = [re.sub(r"<think>.*?</think>", "", a, flags=re.DOTALL).strip() for a in answers]
    else:
        answers = [a.strip() for a in answers]
    return answers


def load_prompts(path):
    """Load prompts as a list of {'id', 'prompt'}.
    .xlsx -> reads the 'qid' and 'prompt' columns; .json -> list of {'id','prompt'}.
    """
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        from openpyxl import load_workbook
        rows = load_workbook(path, read_only=True, data_only=True).active.iter_rows(values_only=True)
        header = list(next(rows))
        qi, pi = header.index("qid"), header.index("prompt")
        return [{"id": r[qi], "prompt": r[pi]} for r in rows if r[pi]]
    if ext == ".json":
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    raise ValueError(f"Unsupported input '{ext}' (use .xlsx or .json)")


def main():
    parser = argparse.ArgumentParser(description="Generate answers with LLMs (Stage 2)")
    parser.add_argument("--model",       required=True, choices=list(MODELS.keys()))
    parser.add_argument("--input",       default="../data/classified/rag_questions.json",
                        help="prompts file: .xlsx (qid+prompt) or .json")
    parser.add_argument("--output-dir",  default="../outputs/answers/")
    parser.add_argument("--max-tokens",  type=int, default=1500)
    parser.add_argument("--batch-size",  type=int, default=16,
                        help="Prompts per generate() call (raise if GPU memory allows)")
    parser.add_argument("--temperature", type=float, default=None,
                        help="Override sampling temperature (default: the model's own; pass 0 for greedy)")
    parser.add_argument("--top-p",       type=float, default=None,
                        help="Override nucleus sampling top-p (default: the model's own)")
    parser.add_argument("--no-quantize", action="store_true", help="Disable 4-bit quantization (for A6000)")
    parser.add_argument("--prompt", choices=list(PROMPTS), default="alt2025-explicit",
                        help="System prompt to use. Non-default writes to <model>_<prompt>.json")
    parser.add_argument("--allow-thinking", action="store_true",
                        help="Let native-thinking models (e.g. Qwen3) reason. <think> blocks are "
                             "split into a separate 'thinking' field; 'answer' keeps the final text.")
    parser.add_argument("--limit", type=int, default=None,
                        help="Only process the first N prompts (for smoke tests)")
    parser.add_argument("--num-shards", type=int, default=1,
                        help="Split prompts across N processes (one per GPU). Each shard "
                             "writes its own .partKofN file; merge them when all finish.")
    parser.add_argument("--shard-id", type=int, default=0,
                        help="Which shard (0..num_shards-1) this process handles.")
    # --- opt-in extensions for the larger-model run (defaults preserve original behaviour) ---
    parser.add_argument("--dtype", choices=["legacy", "native", "fp16", "bf16"], default="legacy",
                        help="legacy = gemma->bf16 else fp16 (original). native = repo's torch_dtype "
                             "(bf16-native models stay bf16).")
    parser.add_argument("--attn", choices=["auto", "sdpa", "eager"], default="auto",
                        help="attn_implementation. sdpa is the fast Turing/sm_75 path.")
    parser.add_argument("--seed", type=int, default=42, help="transformers.set_seed for reproducibility.")
    parser.add_argument("--emit-meta", action="store_true",
                        help="Write a sidecar <model>_<prompt>.meta.json (resolved gen config, dtype, "
                             "attn, seed, peak VRAM) and append to config_table.tsv.")
    parser.add_argument("--print-first-prompt", action="store_true",
                        help="Print the fully rendered prompt string for the first example (leak check).")
    parser.add_argument("--smoke", type=int, default=None,
                        help="Smoke test: generate only the first N prompts, print them, check for "
                             "NaN/empty/degenerate output; exit 0 (clean) or 2 (degenerate). No full run.")
    args = parser.parse_args()
    if not (0 <= args.shard_id < args.num_shards):
        parser.error("--shard-id must be in [0, --num-shards)")

    from transformers import set_seed
    set_seed(args.seed)

    global SYSTEM_PROMPT, ALLOW_THINKING
    SYSTEM_PROMPT = PROMPTS[args.prompt]
    ALLOW_THINKING = args.allow_thinking
    print(f"System prompt: {args.prompt}  |  allow_thinking={ALLOW_THINKING}")

    prompts = load_prompts(args.input)
    if args.limit:
        prompts = prompts[:args.limit]
    print(f"Loaded {len(prompts)} prompts from {args.input}")
    if args.num_shards > 1:                       # data-parallel: 1 shard per GPU process
        prompts = [p for i, p in enumerate(prompts) if i % args.num_shards == args.shard_id]
        print(f"Shard {args.shard_id}/{args.num_shards}: handling {len(prompts)} prompts")

    global GEN_TOP_K, GEN_DO_SAMPLE
    model_id = MODELS[args.model]
    # dtype: --dtype legacy = the original rule (gemma->bf16 else fp16); native = repo dtype
    # (bf16-native models stay bf16 — the safe choice for Gemma soft-capping). See resolve_dtype.
    if args.dtype == "legacy":
        load_dtype = torch.bfloat16 if args.model.startswith("gemma") else torch.float16
    else:
        load_dtype = resolve_dtype(model_id, args.dtype)
    if torch.cuda.is_available():
        torch.cuda.reset_peak_memory_stats()
    print(f"Loading {args.model} ({model_id})  quantize={not args.no_quantize}  dtype={load_dtype}  attn={args.attn}")
    try:
        tokenizer, model = load_model(model_id, quantize=not args.no_quantize, dtype=load_dtype, attn=args.attn)
    except Exception as e:
        if args.attn == "sdpa":                   # some custom archs reject sdpa -> retry eager
            print(f"  [attn] sdpa load failed ({type(e).__name__}: {e}); retrying attn=eager")
            args.attn = "eager"
            tokenizer, model = load_model(model_id, quantize=not args.no_quantize, dtype=load_dtype, attn="eager")
        else:
            raise
    print("Model loaded.")

    gen = resolve_gen_config(model_id, args.model, args.temperature, args.top_p)
    temperature, top_p = gen["temperature"], gen["top_p"]
    GEN_TOP_K, GEN_DO_SAMPLE = gen["top_k"], gen["do_sample"]
    print(f"Resolved generation config (creator-default): {gen}")

    # ---- rendered-prompt preview (verify no system content leaked into the user turn) ----
    def render(p):
        manual = MANUAL_TEMPLATES.get(args.model)
        if manual:
            return manual.format(system=SYSTEM_PROMPT, prompt=p)
        extra = {} if ALLOW_THINKING else THINKING_KWARGS.get(args.model, {})
        return tokenizer.apply_chat_template(build_messages(p, args.model), tokenize=False,
                                             add_generation_prompt=True, **extra)
    if args.print_first_prompt and prompts:
        print("=" * 70 + f"\nFIRST RENDERED PROMPT ({args.model}):\n" + "-" * 70)
        print(render(prompts[0]["prompt"]))
        print("=" * 70)

    # ---- smoke test: first N prompts, flag NaN/empty/degenerate fp16 output, then exit ----
    if args.smoke:
        sm = prompts[:args.smoke]
        ans = generate_with_retry([c["prompt"] for c in sm], tokenizer, model,
                                  args.model, args.max_tokens, temperature, top_p)
        bad = 0
        for i, (c, a) in enumerate(zip(sm, ans)):
            deg = is_degenerate(a); bad += deg
            print(f"\n[smoke {i} · {'DEGENERATE' if deg else 'ok'}] {c['id']}\n{a[:300]}")
        ok = bad == 0
        print(f"\nSMOKE {'PASS' if ok else 'FAIL'} — {bad}/{len(ans)} degenerate for {args.model}")
        sys.exit(0 if ok else 2)

    os.makedirs(args.output_dir, exist_ok=True)
    suffix = "" if args.prompt == "default" else f"_{args.prompt}"
    shard_sfx = f".part{args.shard_id}of{args.num_shards}" if args.num_shards > 1 else ""
    out_path = os.path.join(args.output_dir, f"{args.model}{suffix}{shard_sfx}.json")

    # Resume: skip prompts already processed
    done_ids = set()
    results = []
    if os.path.exists(out_path):
        with open(out_path, encoding="utf-8") as f:
            results = json.load(f)
        before = len(results)
        results = [r for r in results if not str(r["answer"]).startswith("ERROR")]  # retry failed ones
        done_ids = {r["id"] for r in results}
        dropped = before - len(results)
        print(f"Resuming — {len(done_ids)} already done, {len(prompts) - len(done_ids)} remaining"
              + (f" (regenerating {dropped} previous ERROR answers)" if dropped else "") + ".")

    def save():
        # atomic write so a crash mid-save can't corrupt the checkpoint
        tmp = out_path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        os.replace(tmp, out_path)

    todo = [p for p in prompts if p["id"] not in done_ids]
    bs = args.batch_size
    save_every = 128                 # checkpoint cadence, independent of batch size
    last_saved = len(results)
    for i in tqdm(range(0, len(todo), bs), desc=args.model):
        chunk = todo[i:i + bs]
        try:
            answers = generate_with_retry([c["prompt"] for c in chunk], tokenizer, model,
                                          args.model, args.max_tokens, temperature, top_p)
        except Exception as e:
            print(f"\nError on batch at {i}: {e}")
            answers = [f"ERROR: {e}"] * len(chunk)
        for c, a in zip(chunk, answers):
            entry = {"id": c["id"], "prompt": c["prompt"], "answer": a, "model": args.model}
            if "</think>" in a:                  # thinking model: keep reasoning separate
                think, _, ans = a.partition("</think>")
                entry["thinking"] = think.replace("<think>", "").strip()
                entry["answer"] = ans.strip()
            results.append(entry)
        if len(results) - last_saved >= save_every:
            save()
            last_saved = len(results)
    save()                            # final flush (remaining < save_every)

    print(f"Saved {len(results)} answers -> {out_path}")

    # ---- sidecar metadata + aggregate config table (the requested config record) ----
    if args.emit_meta:
        peak = {f"cuda:{d}": round(torch.cuda.max_memory_allocated(d) / 1e9, 2)
                for d in range(torch.cuda.device_count())} if torch.cuda.is_available() else {}
        total_vram = round(sum(peak.values()), 2)
        meta = {
            "model": args.model, "model_id": model_id,
            "dtype": str(load_dtype).replace("torch.", ""), "attn": args.attn, "seed": args.seed,
            "prompt": args.prompt, "allow_thinking": ALLOW_THINKING,
            "gen_config": gen, "max_tokens": args.max_tokens,
            "peak_vram_gb": peak, "peak_vram_total_gb": total_vram, "n_answers": len(results),
        }
        with open(os.path.join(args.output_dir, f"{args.model}{suffix}.meta.json"), "w", encoding="utf-8") as f:
            json.dump(meta, f, ensure_ascii=False, indent=2)
        ct = os.path.join(args.output_dir, "config_table.tsv")
        cols = ["model", "model_id", "engine", "dtype", "attn", "seed", "do_sample", "temperature",
                "top_p", "top_k", "thinking", "max_tokens", "peak_vram_total_gb", "n_answers"]
        row = [args.model, model_id, "hf", meta["dtype"], args.attn, args.seed, gen["do_sample"],
               gen["temperature"], gen["top_p"], gen["top_k"], ALLOW_THINKING, args.max_tokens,
               total_vram, len(results)]
        new = not os.path.exists(ct)
        with open(ct, "a", encoding="utf-8") as f:
            if new: f.write("\t".join(cols) + "\n")
            f.write("\t".join(str(x) for x in row) + "\n")
        print(f"Wrote {args.model}{suffix}.meta.json  (peak VRAM {total_vram} GB)  + config_table.tsv row")


if __name__ == "__main__":
    main()
