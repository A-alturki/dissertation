"""
Prompt frontier LLMs on the sample questions with ALT_PROMPT_2025_explicit.

Reuses annotate_spans.py's multi-provider layer (call_api + MODEL_REGISTRY, .env keys).
The system prompt is read from inference_vllm.py (ALT_PROMPT_2025_explicit) via AST,
so it isn't retyped and vLLM isn't imported. One answer per question per model,
concurrent, with resume. Writes <model>_sample.json next to the other samples.

Usage:
    python frontier_answers.py                       # all 3 models, full sample
    python frontier_answers.py --limit 3 --workers 2 # quick test
    python frontier_answers.py --models gpt-5         # subset
"""
import os, sys, json, time, argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm.auto import tqdm

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from annotate_spans import call_api, MODEL_REGISTRY   # also runs load_dotenv()

DEFAULT_MODELS = ["gemini-3-flash-preview", "gpt-5.4", "gpt-5"]

# Approx USD per 1M tokens (input, output) — PLACEHOLDERS; edit to current pricing.
PRICES = {
    "gemini-3-flash-preview": (0.30, 2.50),
    "gpt-5.4":                (1.25, 10.0),
    "gpt-5":                  (1.25, 10.0),
}

SYSTEM = "أجب عن السؤال التالي و استشهد بآيات من القرآن الكريم و احاديث شريفة عند الاستشهاد بآية قرآنية، اذكر اسم السورة ورقم الآية. عند الاستشهاد بحديث، اذكر المصدر (البخاري، مسلم، إلخ)"



def load_questions(path):
    """Load [{'id','prompt'}] from a sample json/xlsx (any model's sample has all qids)."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".json":
        d = json.load(open(path, encoding="utf-8"))
        return [{"id": x["id"], "prompt": x["prompt"]} for x in d]
    if ext == ".xlsx":
        from openpyxl import load_workbook
        rows = load_workbook(path, read_only=True, data_only=True).active.iter_rows(values_only=True)
        h = list(next(rows)); qi, pi = h.index("qid") if "qid" in h else h.index("id"), h.index("prompt")
        return [{"id": r[qi], "prompt": r[pi]} for r in rows if r[pi]]
    raise SystemExit(f"Unsupported questions file: {path}")


def run_model(model, questions, outdir, workers, save_every):
    out_path = os.path.join(outdir, f"{model}_sample.json")
    results = {}
    if os.path.exists(out_path):
        for r in json.load(open(out_path, encoding="utf-8")):
            results[r["id"]] = r
    todo = [q for q in questions if not str(results.get(q["id"], {}).get("answer", "")).strip()
            or str(results.get(q["id"], {}).get("answer", "")).startswith("ERROR")]

    def save():
        tmp = out_path + ".tmp"
        json.dump([results[q["id"]] for q in questions if q["id"] in results],
                  open(tmp, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
        os.replace(tmp, out_path)

    def one(item):
        t0 = time.time()
        try:
            text, usage = call_api(model, SYSTEM, str(item["prompt"]), max_retries=3, schema=None)
        except Exception as e:
            text, usage = f"ERROR: {str(e)[:150]}", None
        return item["id"], item["prompt"], (text.strip() if text else "ERROR: no response"), usage, time.time() - t0

    print(f"\n=== {model}: {len(todo)}/{len(questions)} to do ===")
    tin = tout = done = 0; lat = []; t_start = time.time()
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futs = [ex.submit(one, q) for q in todo]
        for f in tqdm(as_completed(futs), total=len(futs), desc=model):
            i, p, a, usage, dt = f.result()
            results[i] = {"id": i, "prompt": p, "answer": a, "model": model}
            if usage:
                tin += usage.get("input_tokens") or 0
                tout += usage.get("output_tokens") or 0
            lat.append(dt); done += 1
            if done % save_every == 0:
                save()
    save()
    wall = time.time() - t_start
    errs = sum(1 for q in questions if str(results.get(q["id"], {}).get("answer", "")).startswith("ERROR"))
    pin, pout = PRICES.get(model, (None, None))
    cost = (tin / 1e6 * pin + tout / 1e6 * pout) if pin is not None else None
    mean_lat = round(sum(lat) / len(lat), 2) if lat else 0.0
    print(f"  in {tin:,} | out {tout:,} tok | wall {wall:.1f}s "
          f"({wall/max(len(todo),1):.2f}s/prompt @ {workers}w, mean call {mean_lat}s)"
          + (f" | ~${cost:.3f}" if cost is not None else "")
          + (f" | {errs} ERR" if errs else ""))
    return {"model": model, "n_calls": len(todo), "input_tokens": tin, "output_tokens": tout,
            "wall_s": round(wall, 1), "s_per_prompt": round(wall / max(len(todo), 1), 2),
            "mean_call_s": mean_lat, "errors": errs,
            "est_cost_usd": round(cost, 4) if cost is not None else None}


def main():
    ap = argparse.ArgumentParser(description="Prompt frontier LLMs on the sample questions")
    ap.add_argument("--questions", default="../outputs/answers/samples/allam-7b_sample.json",
                    help="any sample file with the qid+prompt set")
    ap.add_argument("--outdir", default="../outputs/answers/samples")
    ap.add_argument("--models", nargs="+", default=DEFAULT_MODELS)
    ap.add_argument("--workers", type=int, default=6)
    ap.add_argument("--save-every", type=int, default=10)
    ap.add_argument("--limit", type=int, default=None)
    args = ap.parse_args()

    for m in args.models:
        if m not in MODEL_REGISTRY:
            sys.exit(f"Unknown model '{m}'. Known: {list(MODEL_REGISTRY)}")

    questions = load_questions(args.questions)
    if args.limit:
        questions = questions[:args.limit]
    os.makedirs(args.outdir, exist_ok=True)
    print(f"{len(questions)} questions | system prompt: ALT_PROMPT_2025_explicit ({len(SYSTEM)} chars)")
    print(f"models: {args.models}")

    metrics = [run_model(m, questions, args.outdir, args.workers, args.save_every) for m in args.models]

    print("\n=== SUMMARY (this run) ===")
    hdr = f"{'model':26}{'calls':>6}{'in_tok':>10}{'out_tok':>10}{'wall_s':>8}{'s/prompt':>9}{'$est':>8}"
    print(hdr); print("-" * len(hdr))
    for x in metrics:
        cost = f"{x['est_cost_usd']:.3f}" if x["est_cost_usd"] is not None else "n/a"
        print(f"{x['model']:26}{x['n_calls']:>6}{x['input_tokens']:>10,}{x['output_tokens']:>10,}"
              f"{x['wall_s']:>8.1f}{x['s_per_prompt']:>9.2f}{cost:>8}")
    mp = os.path.join(args.outdir, "frontier_run_metrics.json")
    json.dump(metrics, open(mp, "w", encoding="utf-8"), indent=2)
    print(f"\nmetrics -> {mp}   (PRICES are placeholders — edit them for real $)")


if __name__ == "__main__":
    main()
