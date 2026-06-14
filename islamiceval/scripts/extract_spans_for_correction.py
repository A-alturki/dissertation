#!/usr/bin/env python3
"""Scrape Quran + Hadith_Matn citation spans from the annotated sample xlsx review
sheets (outputs/answers/samples/*_sample_annotated.xlsx) into one CSV, for the
correction task. Each row = one citation span.

The `citations` cell of each sheet lists one entry per span, formatted by
build_annotated_xlsx.fmt_span as:
    1. <type>/<label>  «<span_text>»  [<start>:<end>]
where type ∈ {Quran, Hadith_Matn, Hadith_Isnad, Reference} and label ∈
{Correct, Incorrect} for Quran/Hadith_Matn (the corpus-corrected verdict baked
into the xlsx). We keep only Quran + Hadith_Matn here.

Output CSV columns: id, model, span, ref, correctness
  ref: quran | hadith ;  correctness: correct | incorrect
Written UTF-8-with-BOM, csv-quoted, so Arabic spans are byte-exact and never
mangled (no tab/newline substitution); opens cleanly in Excel too.
"""
import argparse, csv, glob, os, re
from openpyxl import load_workbook

# One citation entry. DOTALL so a span_text spanning newlines is captured whole;
# the non-greedy «(.*?)» is pinned by the trailing [start:end] so it can't overrun.
ENTRY = re.compile(
    r"\d+\.\s+(?P<type>Quran|Hadith_Matn|Hadith_Isnad|Reference)"
    r"(?:/(?P<label>Correct|Incorrect))?\s+«(?P<txt>.*?)»\s+\[(?P<s>-?\d+):(?P<e>-?\d+)\]",
    re.DOTALL,
)
REF = {"Quran": "quran", "Hadith_Matn": "hadith"}   # types we keep


def model_of(path):
    return os.path.basename(path).replace("_sample_annotated.xlsx", "")


def rows_from_sheet(path):
    model = model_of(path)
    ws = load_workbook(path, read_only=True)["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c) for c in rows[0]]
    qi, ai, ci = hdr.index("id"), hdr.index("answer"), hdr.index("citations")
    out = []
    for r in rows[1:]:
        qid = r[qi]
        if qid is None or str(qid).startswith("SUMMARY") or qid == "ALL":
            continue                                   # skip the summary block
        cell = r[ci]
        if not cell:
            continue
        answer = str(r[ai] or "")
        for m in ENTRY.finditer(str(cell)):
            t = m.group("type")
            if t not in REF:                           # keep Quran + Hadith_Matn only
                continue
            label = (m.group("label") or "").lower()   # correct | incorrect
            # The «» text is the annotator's ABBREVIATED label ("start .... end").
            # The real span is answer[span_start:span_end] (same slice the corpus
            # verdict uses) — take it verbatim, untouched.
            s, e = int(m.group("s")), int(m.group("e"))
            span = answer[s:e] if 0 <= s < e <= len(answer) else m.group("txt")
            out.append({
                "id": qid,
                "model": model,
                "span": span,
                "ref": REF[t],
                "correctness": label,
            })
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--indir", default="../outputs/answers/samples")
    ap.add_argument("--out", default="../outputs/correction/citation_spans.csv")
    ap.add_argument("--incorrect-only", action="store_true",
                    help="emit only spans whose verdict is incorrect")
    args = ap.parse_args()

    paths = sorted(glob.glob(os.path.join(args.indir, "*_sample_annotated.xlsx")))
    all_rows = []
    for p in paths:
        all_rows.extend(rows_from_sheet(p))
    if args.incorrect_only:
        all_rows = [r for r in all_rows if r["correctness"] == "incorrect"]

    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    with open(args.out, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["id", "model", "span", "ref", "correctness"])
        w.writeheader()
        w.writerows(all_rows)

    # ── summary ──
    from collections import Counter
    by_ref = Counter((r["ref"], r["correctness"]) for r in all_rows)
    print(f"Sheets: {len(paths)}   rows written: {len(all_rows)} -> {args.out}")
    for ref in ("quran", "hadith"):
        c = by_ref.get((ref, "correct"), 0); i = by_ref.get((ref, "incorrect"), 0)
        print(f"  {ref:7s} correct={c:4d}  incorrect={i:4d}  total={c+i}")


if __name__ == "__main__":
    main()
